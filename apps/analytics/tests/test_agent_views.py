"""
Tests for the Agent views with various API actions.
"""

import io

from django.contrib.auth.hashers import make_password
from django.test import Client, TestCase
from django.urls import reverse

import openpyxl

from apps.analytics.models import Agent
from apps.common.constants import SCOPE_CAMPAIGN, SCOPE_NONE
from apps.users.models import Campaign, Role, User


class AgentViewsBaseTestCase(TestCase):
    """Base test case for Agent views with various API actions."""

    fixtures = ["prod/roles.json", "prod/campaigns.json", "prod/users.json"]

    def setUp(self):
        """Set up test data."""
        # Create roles with different permissions
        self.admin_role = Role.objects.create(
            name="TEST ADMIN ROLE",
            is_active=True,
            can_add_agent=True,
            can_edit_agent=True,
            can_delete_agent=True,
            can_export_agent=True,
            can_history_agent=True,
            can_import_agent=True,
            scope_agent=SCOPE_CAMPAIGN,
        )

        self.list_only_role = Role.objects.create(
            name="TEST LIST ONLY ROLE",
            is_active=True,
            can_add_agent=False,
            can_edit_agent=False,
            can_delete_agent=False,
            can_export_agent=False,
            can_history_agent=True,  # This enables can_list_agent
            scope_agent=SCOPE_CAMPAIGN,
        )

        self.no_perm_role = Role.objects.create(
            name="TEST NO PERM ROLE",
            is_active=True,
            can_add_agent=False,
            can_edit_agent=False,
            can_delete_agent=False,
            can_export_agent=False,
            can_history_agent=False,
            scope_agent=SCOPE_NONE,
        )

        # Create a test campaign
        self.test_campaign = Campaign.objects.create(
            name="TEST CAMPAIGN", description="Campaign for testing", is_active=True
        )

        # Create users with different roles
        self.admin_user = User.objects.create(
            username="99999999",
            password=make_password("password"),
            first_name="ADMIN",
            last_name="USER",
            document_type=1,
            document_number="99999999",
            role=self.admin_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.list_only_user = User.objects.create(
            username="88888888",
            password=make_password("password"),
            first_name="LIST ONLY",
            last_name="USER",
            document_type=1,
            document_number="88888888",
            role=self.list_only_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.no_perm_user = User.objects.create(
            username="77777777",
            password=make_password("password"),
            first_name="NO PERM",
            last_name="USER",
            document_type=1,
            document_number="77777777",
            role=self.no_perm_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a test agent that will be the target of actions
        self.target_agent = Agent.objects.create(
            name="TARGET AGENT",
            create_user=self.admin_user,
            modify_user=self.admin_user,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a soft-deleted agent for reactivation tests
        self.deleted_agent = Agent.objects.create(
            name="DELETED AGENT",
            create_user=self.admin_user,
            modify_user=self.admin_user,
            campaign=self.test_campaign,
            is_active=False,
        )

        # Create clients for each user
        self.admin_client = Client()
        self.list_only_client = Client()
        self.no_perm_client = Client()
        self.anonymous_client = Client()

        # Log in the clients
        self.admin_client.login(username="99999999", password="password")
        self.list_only_client.login(username="88888888", password="password")
        self.no_perm_client.login(username="77777777", password="password")


class AgentAddViewTestCase(AgentViewsBaseTestCase):
    """Test case for the Agent views with API_ACTION_ADD action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.add_url = reverse("analytics:agent:add")

    def test_get_add_admin_user(self):
        """User with can_add_agent permission can access the add form."""
        response = self.admin_client.get(self.add_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")

    def test_get_add_list_only_user(self):
        """User with only list permission cannot access the add form."""
        response = self.list_only_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_no_perm_user(self):
        """User with no permissions cannot access the add form."""
        response = self.no_perm_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_anonymous_user(self):
        """Test that an anonymous user cannot access the add form."""
        response = self.anonymous_client.get(self.add_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_add_admin_user_success(self):
        """User with can_add_agent permission can add a new agent."""
        data = {"name": "NEW AGENT"}
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify agent was created
        new_agent = Agent.objects.filter(name="NEW AGENT").first()
        self.assertIsNotNone(new_agent)
        self.assertTrue(new_agent.is_active)

    def test_post_add_admin_user_failure(self):
        """User with can_add_agent permission gets form errors on invalid data."""
        data = {"name": ""}  # Empty name should fail validation
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify agent was not created
        self.assertFalse(Agent.objects.filter(name="").exists())

    def test_post_add_list_only_user_forbidden(self):
        """User with only list permission cannot add a new agent."""
        data = {"name": "SHOULD NOT BE CREATED"}
        response = self.list_only_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify agent was not created
        self.assertFalse(Agent.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_post_add_no_perm_user_forbidden(self):
        """User with no permissions cannot add a new agent."""
        data = {"name": "SHOULD NOT BE CREATED"}
        response = self.no_perm_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify agent was not created
        self.assertFalse(Agent.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_post_add_anonymous_user_redirect(self):
        """Test that an anonymous user cannot add a new agent."""
        data = {"name": "SHOULD NOT BE CREATED"}
        response = self.anonymous_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify agent was not created
        self.assertFalse(Agent.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_delete_add_bad_request(self):
        """Test that DELETE method is not allowed for add."""
        response = self.admin_client.delete(self.add_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class AgentHomeViewTestCase(AgentViewsBaseTestCase):
    """Test case for the Agent views with API_ACTION_HOME action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.home_url = reverse("analytics:agent:home")

    def test_get_home_admin_user(self):
        """User with can_list_agent permission can access the home page."""
        response = self.admin_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        # Validate template used
        self.assertTemplateUsed(response, "analytics/agent/home.html")

    def test_get_home_list_only_user(self):
        """User with only list permission can access the home page."""
        response = self.list_only_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "analytics/agent/home.html")

    def test_get_home_no_perm_user(self):
        """User with no permissions cannot access the home page."""
        response = self.no_perm_client.get(self.home_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_home_anonymous_user(self):
        """Test that an anonymous user cannot access the home page."""
        response = self.anonymous_client.get(self.home_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_home_bad_request(self):
        """Test that POST method is not allowed for home page."""
        response = self.admin_client.post(self.home_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class AgentEditViewTestCase(AgentViewsBaseTestCase):
    """Test case for the Agent views with API_ACTION_EDIT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.edit_url = reverse("analytics:agent:edit", args=[self.target_agent.pk])

    def test_get_edit_admin_user(self):
        """User with can_edit_agent permission can access the edit form."""
        response = self.admin_client.get(self.edit_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, self.target_agent.name)

    def test_get_edit_list_only_user(self):
        """User with only list permission cannot access the edit form."""
        response = self.list_only_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_no_perm_user(self):
        """User with no permissions cannot access the edit form."""
        response = self.no_perm_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_anonymous_user(self):
        """Test that an anonymous user cannot access the edit form."""
        response = self.anonymous_client.get(self.edit_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_edit_admin_user_success(self):
        """User with can_edit_agent permission can edit an agent."""
        data = {"name": "UPDATED AGENT"}
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.target_agent.refresh_from_db()
        self.assertEqual(self.target_agent.name, "UPDATED AGENT")

    def test_post_edit_admin_user_failure(self):
        """User with can_edit_agent permission gets form errors on invalid data."""
        data = {"name": ""}  # Empty name should fail validation
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Refresh from database to verify no changes
        self.target_agent.refresh_from_db()
        self.assertNotEqual(self.target_agent.name, "")

    def test_post_edit_list_only_user_forbidden(self):
        """User with only list permission cannot edit an agent."""
        data = {"name": "SHOULD NOT UPDATE"}
        response = self.list_only_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.target_agent.refresh_from_db()
        self.assertNotEqual(self.target_agent.name, "SHOULD NOT UPDATE")

    def test_post_edit_no_perm_user_forbidden(self):
        """User with no permissions cannot edit an agent."""
        data = {"name": "SHOULD NOT UPDATE"}
        response = self.no_perm_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.target_agent.refresh_from_db()
        self.assertNotEqual(self.target_agent.name, "SHOULD NOT UPDATE")

    def test_post_edit_anonymous_user_redirect(self):
        """Test that an anonymous user cannot edit an agent."""
        data = {"name": "SHOULD NOT UPDATE"}
        response = self.anonymous_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.target_agent.refresh_from_db()
        self.assertNotEqual(self.target_agent.name, "SHOULD NOT UPDATE")

    def test_delete_edit_bad_request(self):
        """Test that DELETE method is not allowed for edit."""
        response = self.admin_client.delete(self.edit_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class AgentDeleteViewTestCase(AgentViewsBaseTestCase):
    """Test case for the Agent views with API_ACTION_DELETE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.delete_url = reverse("analytics:agent:delete", args=[self.target_agent.pk])

    def test_delete_admin_user_success(self):
        """User with can_delete_agent permission can soft-delete an agent."""
        response = self.admin_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.target_agent.refresh_from_db()
        self.assertFalse(self.target_agent.is_active)

    def test_delete_list_only_user_forbidden(self):
        """User with only list permission cannot soft-delete an agent."""
        response = self.list_only_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.target_agent.refresh_from_db()
        self.assertTrue(self.target_agent.is_active)

    def test_delete_no_perm_user_forbidden(self):
        """User with no permissions cannot soft-delete an agent."""
        response = self.no_perm_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.target_agent.refresh_from_db()
        self.assertTrue(self.target_agent.is_active)

    def test_delete_anonymous_user_redirect(self):
        """Test that an anonymous user cannot soft-delete an agent."""
        response = self.anonymous_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.target_agent.refresh_from_db()
        self.assertTrue(self.target_agent.is_active)

    def test_get_delete_bad_request(self):
        """Test that GET method is not allowed for delete."""
        response = self.admin_client.get(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_post_delete_bad_request(self):
        """Test that POST method is not allowed for delete."""
        response = self.admin_client.post(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class AgentReactivateViewTestCase(AgentViewsBaseTestCase):
    """Test case for the Agent views with API_ACTION_REACTIVATE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.reactivate_url = reverse("analytics:agent:reactivate", args=[self.deleted_agent.pk])

    def test_post_reactivate_admin_user_success(self):
        """User with can_delete_agent permission can reactivate a soft-deleted agent."""
        response = self.admin_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.deleted_agent.refresh_from_db()
        self.assertTrue(self.deleted_agent.is_active)

    def test_post_reactivate_list_only_user_forbidden(self):
        """User with only list permission cannot reactivate a soft-deleted agent."""
        response = self.list_only_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.deleted_agent.refresh_from_db()
        self.assertFalse(self.deleted_agent.is_active)

    def test_post_reactivate_no_perm_user_forbidden(self):
        """User with no permissions cannot reactivate a soft-deleted agent."""
        response = self.no_perm_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.deleted_agent.refresh_from_db()
        self.assertFalse(self.deleted_agent.is_active)

    def test_post_reactivate_anonymous_user_redirect(self):
        """Test that an anonymous user cannot reactivate a soft-deleted agent."""
        response = self.anonymous_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.deleted_agent.refresh_from_db()
        self.assertFalse(self.deleted_agent.is_active)

    def test_get_reactivate_bad_request(self):
        """Test that GET method is not allowed for reactivate."""
        response = self.admin_client.get(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_reactivate_bad_request(self):
        """Test that DELETE method is not allowed for reactivate."""
        response = self.admin_client.delete(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class AgentReadViewTestCase(AgentViewsBaseTestCase):
    """Test case for the Agent views with API_ACTION_READ action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.read_url = reverse("analytics:agent:read", args=[self.target_agent.pk])

    def test_get_read_admin_user(self):
        """User with can_list_agent permission can read an agent."""
        response = self.admin_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_agent.name)
        # Check that form fields are disabled
        self.assertContains(response, "disabled")

    def test_get_read_list_only_user(self):
        """User with only list permission can read an agent."""
        response = self.list_only_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_agent.name)
        # Check that form fields are disabled
        self.assertContains(response, "disabled")

    def test_get_read_no_perm_user_forbidden(self):
        """User with no permissions cannot read an agent."""
        response = self.no_perm_client.get(self.read_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_read_anonymous_user_redirect(self):
        """Test that an anonymous user cannot read an agent."""
        response = self.anonymous_client.get(self.read_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_read_bad_request(self):
        """Test that POST method is not allowed for read."""
        response = self.admin_client.post(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_read_bad_request(self):
        """Test that DELETE method is not allowed for read."""
        response = self.admin_client.delete(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class AgentHistoryViewTestCase(AgentViewsBaseTestCase):
    """Test case for the Agent views with API_ACTION_HISTORY action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.history_url = reverse("analytics:agent:history", args=[self.target_agent.pk])

    def test_get_history_admin_user(self):
        """User with can_history_agent permission can view history of an agent."""
        response = self.admin_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion

    def test_get_history_list_only_user(self):
        """User with only list permission can view history of an agent."""
        response = self.list_only_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion

    def test_get_history_no_perm_user_forbidden(self):
        """User with no permissions cannot view history of an agent."""
        response = self.no_perm_client.get(self.history_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_history_anonymous_user_redirect(self):
        """Test that an anonymous user cannot view history of an agent."""
        response = self.anonymous_client.get(self.history_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_history_bad_request(self):
        """Test that POST method is not allowed for history."""
        response = self.admin_client.post(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_history_bad_request(self):
        """Test that DELETE method is not allowed for history."""
        response = self.admin_client.delete(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class AgentListViewTestCase(AgentViewsBaseTestCase):
    """Test case for the Agent views with API_ACTION_LIST action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.list_url = reverse("analytics:agent:list")

    def test_get_list_admin_user(self):
        """User with can_list_agent permission can list agents."""
        response = self.admin_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the response contains a table
        self.assertContains(response, "<table")
        # Check that agents are listed
        self.assertContains(response, self.target_agent.name)
        self.assertContains(response, self.deleted_agent.name)

    def test_get_list_list_only_user(self):
        """User with only list permission can list agents."""
        response = self.list_only_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the response contains a table
        self.assertContains(response, "<table")
        # Check that agents are listed
        self.assertContains(response, self.target_agent.name)
        self.assertContains(response, self.deleted_agent.name)

    def test_get_list_no_perm_user_forbidden(self):
        """User with no permissions cannot list agents."""
        response = self.no_perm_client.get(self.list_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_list_anonymous_user_redirect(self):
        """Test that an anonymous user cannot list agents."""
        response = self.anonymous_client.get(self.list_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_list_bad_request(self):
        """Test that POST method is not allowed for list."""
        response = self.admin_client.post(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_list_bad_request(self):
        """Test that DELETE method is not allowed for list."""
        response = self.admin_client.delete(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class AgentExportViewTestCase(AgentViewsBaseTestCase):
    """Test case for the Agent views with API_ACTION_EXPORT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.export_url = reverse("analytics:agent:export")

    def test_get_export_admin_user(self):
        """User with can_export_agent permission can export agents."""
        response = self.admin_client.get(self.export_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment; filename=", response.get("Content-Disposition", ""))

        # check for items in excel file
        try:
            file_like_object = io.BytesIO(response.content)
            workbook = openpyxl.load_workbook(file_like_object)
        except Exception as e:
            self.fail(f"Error reading excel file: {e}")
        else:
            sheet = workbook.active
            self.assertGreater(sheet.max_row, 1, "No data in excel file")
            agents_id = list()
            for row_index in range(2, sheet.max_row + 1):
                row = sheet[row_index]
                agents_id.append(row[0].value)  # row[0] first column => id column
            self.assertTrue(Agent.objects.filter(pk__in=agents_id).exists())

    def test_get_export_list_only_user_forbidden(self):
        """User with only list permission cannot export agents."""
        response = self.list_only_client.get(self.export_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_no_perm_user_forbidden(self):
        """User with no permissions cannot export agents."""
        response = self.no_perm_client.get(self.export_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_anonymous_user_redirect(self):
        """Test that an anonymous user cannot export agents."""
        response = self.anonymous_client.get(self.export_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_export_bad_request(self):
        """Test that POST method is not allowed for export."""
        response = self.admin_client.post(self.export_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class AgentImportViewTestCase(AgentViewsBaseTestCase):
    """Test case for the Agent views with API_ACTION_IMPORT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.import_url = reverse("analytics:agent:import")

    def test_get_import_admin_user(self):
        """User with can_import_agent permission can access the import form."""
        response = self.admin_client.get(self.import_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, "Instrucciones para el archivo Excel")  # Check instructions
        self.assertContains(response, "name")  # Check for column name

    def test_get_import_list_only_user(self):
        """User with only list permission cannot access the import form."""
        response = self.list_only_client.get(self.import_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_import_no_perm_user(self):
        """User with no permissions cannot access the import form."""
        response = self.no_perm_client.get(self.import_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_import_anonymous_user(self):
        """Test that an anonymous user cannot access the import form."""
        response = self.anonymous_client.get(self.import_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_import_admin_user_success(self):
        """User with can_import_agent permission can import agents."""
        # Create a test Excel file with agent data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["name"])  # Header row
        sheet.append(["IMPORTED AGENT 1"])
        sheet.append(["IMPORTED AGENT 2"])

        # Save the workbook to a BytesIO object
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        # Create a file object for the request
        file_data = {"file": excel_file}

        response = self.admin_client.post(self.import_url, file_data)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Verify agents were created
        self.assertTrue(Agent.objects.filter(name="IMPORTED AGENT 1").exists())
        self.assertTrue(Agent.objects.filter(name="IMPORTED AGENT 2").exists())

        # Verify agents belong to the admin user's campaign
        imported_agent1 = Agent.objects.get(name="IMPORTED AGENT 1")
        self.assertEqual(imported_agent1.campaign, self.admin_user.campaign)

        imported_agent2 = Agent.objects.get(name="IMPORTED AGENT 2")
        self.assertEqual(imported_agent2.campaign, self.admin_user.campaign)

    def test_post_import_list_only_user_forbidden(self):
        """User with only list permission cannot import agents."""
        # Create a test Excel file with agent data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["name"])  # Header row
        sheet.append(["SHOULD NOT BE IMPORTED"])

        # Save the workbook to a BytesIO object
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        # Create a file object for the request
        file_data = {"file": excel_file}

        response = self.list_only_client.post(self.import_url, file_data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify agent was not created
        self.assertFalse(Agent.objects.filter(name="SHOULD NOT BE IMPORTED").exists())

    def test_post_import_no_perm_user_forbidden(self):
        """User with no permissions cannot import agents."""
        # Create a test Excel file with agent data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["name"])  # Header row
        sheet.append(["SHOULD NOT BE IMPORTED"])

        # Save the workbook to a BytesIO object
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        # Create a file object for the request
        file_data = {"file": excel_file}

        response = self.no_perm_client.post(self.import_url, file_data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify agent was not created
        self.assertFalse(Agent.objects.filter(name="SHOULD NOT BE IMPORTED").exists())

    def test_post_import_anonymous_user_redirect(self):
        """Test that an anonymous user cannot import agents."""
        # Create a test Excel file with agent data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["name"])  # Header row
        sheet.append(["SHOULD NOT BE IMPORTED"])

        # Save the workbook to a BytesIO object
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        # Create a file object for the request
        file_data = {"file": excel_file}

        response = self.anonymous_client.post(self.import_url, file_data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify agent was not created
        self.assertFalse(Agent.objects.filter(name="SHOULD NOT BE IMPORTED").exists())

    def test_delete_import_bad_request(self):
        """Test that DELETE method is not allowed for import."""
        response = self.admin_client.delete(self.import_url)
        self.assertEqual(response.status_code, 400)  # Bad request method
