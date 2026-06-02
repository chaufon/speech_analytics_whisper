import io

from django.contrib.auth.hashers import make_password
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.urls import reverse

import openpyxl

from apps.analytics.models import Pattern, Typification
from apps.common.constants import SCOPE_CAMPAIGN, SCOPE_NONE
from apps.users.models import Campaign, Role, User


class TypificationViewsBaseTestCase(TestCase):
    """Base test case for Typification views with various API actions."""

    fixtures = ["prod/roles.json", "prod/campaigns.json", "prod/users.json"]

    def setUp(self):
        """Set up test data."""
        # Create roles with different permissions
        self.admin_role = Role.objects.create(
            name="TEST ADMIN ROLE",
            is_active=True,
            can_add_typification=True,
            can_edit_typification=True,
            can_delete_typification=True,
            can_export_typification=True,
            can_history_typification=True,
            can_import_typification=True,
            scope_typification=SCOPE_CAMPAIGN,
        )

        self.list_only_role = Role.objects.create(
            name="TEST LIST ONLY ROLE",
            is_active=True,
            can_add_typification=False,
            can_edit_typification=False,
            can_delete_typification=False,
            can_export_typification=False,
            can_history_typification=True,
            scope_typification=SCOPE_CAMPAIGN,
        )

        self.no_perm_role = Role.objects.create(
            name="TEST NO PERM ROLE",
            is_active=True,
            can_add_typification=False,
            can_edit_typification=False,
            can_delete_typification=False,
            can_export_typification=False,
            can_history_typification=False,
            scope_typification=SCOPE_NONE,
        )

        # Create a test campaign
        self.test_campaign = Campaign.objects.create(
            name="TEST CAMPAIGN", description="Campaign for testing", is_active=True
        )

        # Create users with different roles
        self.admin_user = User.objects.create(
            username="99999999",
            password=make_password("password"),
            first_name="ADMIN",
            last_name="USER",
            document_type=1,
            document_number="99999999",
            role=self.admin_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.list_only_user = User.objects.create(
            username="88888888",
            password=make_password("password"),
            first_name="LIST ONLY",
            last_name="USER",
            document_type=1,
            document_number="88888888",
            role=self.list_only_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.no_perm_user = User.objects.create(
            username="77777777",
            password=make_password("password"),
            first_name="NO PERM",
            last_name="USER",
            document_type=1,
            document_number="77777777",
            role=self.no_perm_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a test typification that will be the target of actions
        self.target_typification = Typification.objects.create(
            name="TARGET TYPIFICATION",
            create_user=self.admin_user,
            modify_user=self.admin_user,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a test pattern related to the target typification
        self.target_pattern = Pattern.objects.create(
            sentence="TARGET PATTERN",
            typification=self.target_typification,
            create_user=self.admin_user,
            modify_user=self.admin_user,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a soft-deleted typification for reactivation tests
        self.deleted_typification = Typification.objects.create(
            name="DELETED TYPIFICATION",
            create_user=self.admin_user,
            modify_user=self.admin_user,
            campaign=self.test_campaign,
            is_active=False,
        )

        # Create a soft-deleted pattern for reactivation tests
        self.deleted_pattern = Pattern.objects.create(
            sentence="DELETED PATTERN",
            typification=self.target_typification,
            create_user=self.admin_user,
            modify_user=self.admin_user,
            campaign=self.test_campaign,
            is_active=False,
        )

        # Create clients for each user
        self.admin_client = Client()
        self.list_only_client = Client()
        self.no_perm_client = Client()
        self.anonymous_client = Client()

        # Log in the clients
        self.admin_client.login(username="99999999", password="password")
        self.list_only_client.login(username="88888888", password="password")
        self.no_perm_client.login(username="77777777", password="password")


class TypificationAddViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Typification views with API_ACTION_ADD action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.add_url = reverse("analytics:typification:add")

    def test_get_add_admin_user(self):
        """User with can_add_typification permission can access the add form."""
        response = self.admin_client.get(self.add_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")

    def test_get_add_list_only_user(self):
        """User with only list permission cannot access the add form."""
        response = self.list_only_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_no_perm_user(self):
        """User with no permissions cannot access the add form."""
        response = self.no_perm_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_anonymous_user(self):
        """Test that an anonymous user cannot access the add form."""
        response = self.anonymous_client.get(self.add_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_add_admin_user_success(self):
        """User with can_add_typification permission can add a new typification."""
        data = {"name": "NEW TYPIFICATION"}
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify typification was created
        new_typification = Typification.objects.filter(name="NEW TYPIFICATION").first()
        self.assertIsNotNone(new_typification)
        self.assertTrue(new_typification.is_active)

    def test_post_add_admin_user_failure(self):
        """User with can_add_typification permission gets form errors on invalid data."""
        data = {"name": ""}  # Empty name should fail validation
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify typification was not created
        self.assertFalse(Typification.objects.filter(name="").exists())

    def test_post_add_list_only_user_forbidden(self):
        """User with only list permission cannot add a new typification."""
        data = {"name": "SHOULD NOT BE CREATED"}
        response = self.list_only_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify typification was not created
        self.assertFalse(Typification.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_post_add_no_perm_user_forbidden(self):
        """User with no permissions cannot add a new typification."""
        data = {"name": "SHOULD NOT BE CREATED"}
        response = self.no_perm_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify typification was not created
        self.assertFalse(Typification.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_post_add_anonymous_user_redirect(self):
        """Test that an anonymous user cannot add a new typification."""
        data = {"name": "SHOULD NOT BE CREATED"}
        response = self.anonymous_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify typification was not created
        self.assertFalse(Typification.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_delete_add_bad_request(self):
        """Test that DELETE method is not allowed for add."""
        response = self.admin_client.delete(self.add_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class TypificationHomeViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Typification views with API_ACTION_HOME action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.home_url = reverse("analytics:typification:home")

    def test_get_home_admin_user(self):
        """User with can_list_typification permission can access the home page."""
        response = self.admin_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        # Validate template used
        self.assertTemplateUsed(response, "analytics/typification/home.html")

    def test_get_home_list_only_user(self):
        """User with only list permission can access the home page."""
        response = self.list_only_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "analytics/typification/home.html")

    def test_get_home_no_perm_user(self):
        """User with no permissions cannot access the home page."""
        response = self.no_perm_client.get(self.home_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_home_anonymous_user(self):
        """Test that an anonymous user cannot access the home page."""
        response = self.anonymous_client.get(self.home_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_home_bad_request(self):
        """Test that POST method is not allowed for home."""
        response = self.admin_client.post(self.home_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class TypificationEditViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Typification views with API_ACTION_EDIT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.edit_url = reverse("analytics:typification:edit", args=[self.target_typification.pk])

    def test_get_edit_admin_user(self):
        """User with can_edit_typification permission can access the edit form."""
        response = self.admin_client.get(self.edit_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, self.target_typification.name)

    def test_get_edit_list_only_user(self):
        """User with only list permission cannot access the edit form."""
        response = self.list_only_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_no_perm_user(self):
        """User with no permissions cannot access the edit form."""
        response = self.no_perm_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_anonymous_user(self):
        """Test that an anonymous user cannot access the edit form."""
        response = self.anonymous_client.get(self.edit_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_edit_admin_user_success(self):
        """User with can_edit_typification permission can edit a typification."""
        data = {"name": "UPDATED TYPIFICATION"}
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify typification was updated
        self.target_typification.refresh_from_db()
        self.assertEqual(self.target_typification.name, "UPDATED TYPIFICATION")

    def test_post_edit_admin_user_failure(self):
        """User with can_edit_typification permission gets form errors on invalid data."""
        data = {"name": ""}  # Empty name should fail validation
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify typification was not updated
        self.target_typification.refresh_from_db()
        self.assertEqual(self.target_typification.name, "TARGET TYPIFICATION")

    def test_post_edit_list_only_user_forbidden(self):
        """User with only list permission cannot edit a typification."""
        data = {"name": "SHOULD NOT BE UPDATED"}
        response = self.list_only_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify typification was not updated
        self.target_typification.refresh_from_db()
        self.assertEqual(self.target_typification.name, "TARGET TYPIFICATION")

    def test_post_edit_no_perm_user_forbidden(self):
        """User with no permissions cannot edit a typification."""
        data = {"name": "SHOULD NOT BE UPDATED"}
        response = self.no_perm_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify typification was not updated
        self.target_typification.refresh_from_db()
        self.assertEqual(self.target_typification.name, "TARGET TYPIFICATION")

    def test_post_edit_anonymous_user_redirect(self):
        """Test that an anonymous user cannot edit a typification."""
        data = {"name": "SHOULD NOT BE UPDATED"}
        response = self.anonymous_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify typification was not updated
        self.target_typification.refresh_from_db()
        self.assertEqual(self.target_typification.name, "TARGET TYPIFICATION")

    def test_delete_edit_bad_request(self):
        """Test that DELETE method is not allowed for edit."""
        response = self.admin_client.delete(self.edit_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class TypificationDeleteViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Typification views with API_ACTION_DELETE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.delete_url = reverse(
            "analytics:typification:delete", args=[self.target_typification.pk]
        )

    def test_delete_admin_user_success(self):
        """User with can_delete_typification permission can delete a typification."""
        response = self.admin_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify typification was soft-deleted
        self.target_typification.refresh_from_db()
        self.assertFalse(self.target_typification.is_active)

        # Verify related patterns were also soft-deleted
        self.target_pattern.refresh_from_db()
        self.assertFalse(self.target_pattern.is_active)

    def test_delete_list_only_user_forbidden(self):
        """User with only list permission cannot delete a typification."""
        response = self.list_only_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify typification was not deleted
        self.target_typification.refresh_from_db()
        self.assertTrue(self.target_typification.is_active)

        # Verify related patterns were not deleted
        self.target_pattern.refresh_from_db()
        self.assertTrue(self.target_pattern.is_active)

    def test_delete_no_perm_user_forbidden(self):
        """User with no permissions cannot delete a typification."""
        response = self.no_perm_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify typification was not deleted
        self.target_typification.refresh_from_db()
        self.assertTrue(self.target_typification.is_active)

        # Verify related patterns were not deleted
        self.target_pattern.refresh_from_db()
        self.assertTrue(self.target_pattern.is_active)

    def test_delete_anonymous_user_redirect(self):
        """Test that an anonymous user cannot delete a typification."""
        response = self.anonymous_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify typification was not deleted
        self.target_typification.refresh_from_db()
        self.assertTrue(self.target_typification.is_active)

        # Verify related patterns were not deleted
        self.target_pattern.refresh_from_db()
        self.assertTrue(self.target_pattern.is_active)

    def test_get_delete_bad_request(self):
        """Test that GET method is not allowed for delete."""
        response = self.admin_client.get(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_post_delete_bad_request(self):
        """Test that POST method is not allowed for delete."""
        response = self.admin_client.post(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class TypificationReactivateViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Typification views with API_ACTION_REACTIVATE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.reactivate_url = reverse(
            "analytics:typification:reactivate", args=[self.deleted_typification.pk]
        )

    def test_post_reactivate_admin_user_success(self):
        """User with can_delete_typification permission can reactivate a typification."""
        response = self.admin_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify typification was reactivated
        self.deleted_typification.refresh_from_db()
        self.assertTrue(self.deleted_typification.is_active)

    def test_post_reactivate_list_only_user_forbidden(self):
        """User with only list permission cannot reactivate a typification."""
        response = self.list_only_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify typification was not reactivated
        self.deleted_typification.refresh_from_db()
        self.assertFalse(self.deleted_typification.is_active)

    def test_post_reactivate_no_perm_user_forbidden(self):
        """User with no permissions cannot reactivate a typification."""
        response = self.no_perm_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify typification was not reactivated
        self.deleted_typification.refresh_from_db()
        self.assertFalse(self.deleted_typification.is_active)

    def test_post_reactivate_anonymous_user_redirect(self):
        """Test that an anonymous user cannot reactivate a typification."""
        response = self.anonymous_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify typification was not reactivated
        self.deleted_typification.refresh_from_db()
        self.assertFalse(self.deleted_typification.is_active)

    def test_get_reactivate_bad_request(self):
        """Test that GET method is not allowed for reactivate."""
        response = self.admin_client.get(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_reactivate_bad_request(self):
        """Test that DELETE method is not allowed for reactivate."""
        response = self.admin_client.delete(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class TypificationReadViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Typification views with API_ACTION_READ action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.read_url = reverse("analytics:typification:read", args=[self.target_typification.pk])

    def test_get_read_admin_user(self):
        """User with can_list_typification permission can access the read view."""
        response = self.admin_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_typification.name)
        # Check that the form is read-only (no submit button)
        self.assertContains(response, "disabled")

    def test_get_read_list_only_user(self):
        """User with only list permission can access the read view."""
        response = self.list_only_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_typification.name)
        # Check that the form is read-only (no submit button)
        self.assertContains(response, "disabled")

    def test_get_read_no_perm_user_forbidden(self):
        """User with no permissions cannot access the read view."""
        response = self.no_perm_client.get(self.read_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_read_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the read view."""
        response = self.anonymous_client.get(self.read_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_read_bad_request(self):
        """Test that POST method is not allowed for read."""
        response = self.admin_client.post(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_read_bad_request(self):
        """Test that DELETE method is not allowed for read."""
        response = self.admin_client.delete(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class TypificationHistoryViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Typification views with API_ACTION_HISTORY action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.history_url = reverse(
            "analytics:typification:history", args=[self.target_typification.pk]
        )

    def test_get_history_admin_user(self):
        """User with can_history_typification permission can access the history view."""
        response = self.admin_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion component

    def test_get_history_list_only_user(self):
        """User with only list permission can access the history view."""
        response = self.list_only_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion component

    def test_get_history_no_perm_user_forbidden(self):
        """User with no permissions cannot access the history view."""
        response = self.no_perm_client.get(self.history_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_history_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the history view."""
        response = self.anonymous_client.get(self.history_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_history_bad_request(self):
        """Test that POST method is not allowed for history."""
        response = self.admin_client.post(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_history_bad_request(self):
        """Test that DELETE method is not allowed for history."""
        response = self.admin_client.delete(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class TypificationListViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Typification views with API_ACTION_LIST action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.list_url = reverse("analytics:typification:list")

    def test_get_list_admin_user(self):
        """User with can_list_typification permission can access the list view."""
        response = self.admin_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the table contains the target typification
        self.assertContains(response, self.target_typification.name)
        # Check that the table contains the deleted typification
        self.assertContains(response, self.deleted_typification.name)

    def test_get_list_list_only_user(self):
        """User with only list permission can access the list view."""
        response = self.list_only_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the table contains the target typification
        self.assertContains(response, self.target_typification.name)
        # Check that the table contains the deleted typification
        self.assertContains(response, self.deleted_typification.name)

    def test_get_list_no_perm_user_forbidden(self):
        """User with no permissions cannot access the list view."""
        response = self.no_perm_client.get(self.list_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_list_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the list view."""
        response = self.anonymous_client.get(self.list_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_list_bad_request(self):
        """Test that POST method is not allowed for list."""
        response = self.admin_client.post(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_list_bad_request(self):
        """Test that DELETE method is not allowed for list."""
        response = self.admin_client.delete(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class TypificationExportIndividualViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Typification views with API_ACTION_EXPORT_INDIVIDUAL action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.export_individual_url = reverse(
            "analytics:typification:export_individual", args=[self.target_typification.pk]
        )

    def test_get_export_individual_admin_user(self):
        """User with can_export_typification permission can export a typification."""
        response = self.admin_client.get(self.export_individual_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(response.has_header("Content-Disposition"))

        # Verify the Excel file contains the expected data
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = wb.active

        # Check that the header row contains "Oración" column
        header_row = [cell.value for cell in sheet[1]]
        self.assertIn("Oración", header_row)

        # Check that the data rows contain the pattern
        sentence_column_index = header_row.index("Oración") + 1  # +1 because openpyxl is 1-indexed
        sentences = [
            sheet.cell(row=i, column=sentence_column_index).value
            for i in range(2, sheet.max_row + 1)
        ]
        self.assertIn(self.target_pattern.sentence, sentences)

    def test_get_export_individual_list_only_user_forbidden(self):
        """User with only list permission cannot export a typification."""
        response = self.list_only_client.get(self.export_individual_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_individual_no_perm_user_forbidden(self):
        """User with no permissions cannot export a typification."""
        response = self.no_perm_client.get(self.export_individual_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_individual_anonymous_user_redirect(self):
        """Test that an anonymous user cannot export a typification."""
        response = self.anonymous_client.get(self.export_individual_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_export_individual_bad_request(self):
        """Test that POST method is not allowed for export_individual."""
        response = self.admin_client.post(self.export_individual_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_export_individual_bad_request(self):
        """Test that DELETE method is not allowed for export_individual."""
        response = self.admin_client.delete(self.export_individual_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class TypificationImportViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Typification views with API_ACTION_IMPORT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.import_url = reverse("analytics:typification:import")

    def test_get_import_admin_user(self):
        """User with can_import_typification permission can access the import form."""
        response = self.admin_client.get(self.import_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, "file")  # Check for file upload field

    def test_get_import_list_only_user(self):
        """User with only list permission cannot access the import form."""
        response = self.list_only_client.get(self.import_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_import_no_perm_user(self):
        """User with no permissions cannot access the import form."""
        response = self.no_perm_client.get(self.import_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_import_anonymous_user(self):
        """Test that an anonymous user cannot access the import form."""
        response = self.anonymous_client.get(self.import_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_import_admin_user_success(self):
        """User with can_import_typification permission can import a typification."""
        # Create a test Excel file
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["sentence"])  # Header row
        sheet.append(["IMPORTED PATTERN 1"])
        sheet.append(["IMPORTED PATTERN 2"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Submit the form with the Excel file
        data = {
            "name": "IMPORTED TYPIFICATION",
            "file": SimpleUploadedFile(
                "test.xlsx",
                excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = self.admin_client.post(self.import_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify typification was created
        imported_typification = Typification.objects.filter(name="IMPORTED TYPIFICATION").first()
        self.assertIsNotNone(imported_typification)
        self.assertTrue(imported_typification.is_active)

        # Verify patterns were created
        imported_patterns = Pattern.objects.filter(typification=imported_typification)
        self.assertEqual(imported_patterns.count(), 2)
        self.assertIn("IMPORTED PATTERN 1", [pattern.sentence for pattern in imported_patterns])
        self.assertIn("IMPORTED PATTERN 2", [pattern.sentence for pattern in imported_patterns])

    def test_post_import_list_only_user_forbidden(self):
        """User with only list permission cannot import a typification."""
        # Create a test Excel file
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["sentence"])  # Header row
        sheet.append(["SHOULD NOT BE IMPORTED"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Submit the form with the Excel file
        data = {
            "name": "SHOULD NOT BE IMPORTED",
            "file": SimpleUploadedFile(
                "test.xlsx",
                excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = self.list_only_client.post(self.import_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify typification was not created
        self.assertFalse(Typification.objects.filter(name="SHOULD NOT BE IMPORTED").exists())

    def test_post_import_no_perm_user_forbidden(self):
        """User with no permissions cannot import a typification."""
        # Create a test Excel file
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["sentence"])  # Header row
        sheet.append(["SHOULD NOT BE IMPORTED"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Submit the form with the Excel file
        data = {
            "name": "SHOULD NOT BE IMPORTED",
            "file": SimpleUploadedFile(
                "test.xlsx",
                excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = self.no_perm_client.post(self.import_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify typification was not created
        self.assertFalse(Typification.objects.filter(name="SHOULD NOT BE IMPORTED").exists())

    def test_post_import_anonymous_user_redirect(self):
        """Test that an anonymous user cannot import a typification."""
        # Create a test Excel file
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["sentence"])  # Header row
        sheet.append(["SHOULD NOT BE IMPORTED"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Submit the form with the Excel file
        data = {
            "name": "SHOULD NOT BE IMPORTED",
            "file": SimpleUploadedFile(
                "test.xlsx",
                excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = self.anonymous_client.post(self.import_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify typification was not created
        self.assertFalse(Typification.objects.filter(name="SHOULD NOT BE IMPORTED").exists())

    def test_delete_import_bad_request(self):
        """Test that DELETE method is not allowed for import."""
        response = self.admin_client.delete(self.import_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class PatternRelatedAddViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Pattern related views with API_ACTION_ADD action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.add_url = reverse(
            "analytics:typification:pattern:add", args=[self.target_typification.pk]
        )

    def test_get_add_admin_user(self):
        """User with can_add_typification permission can access the add form."""
        response = self.admin_client.get(self.add_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")

    def test_get_add_list_only_user(self):
        """User with only list permission cannot access the add form."""
        response = self.list_only_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_no_perm_user(self):
        """User with no permissions cannot access the add form."""
        response = self.no_perm_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_anonymous_user(self):
        """Test that an anonymous user cannot access the add form."""
        response = self.anonymous_client.get(self.add_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_add_admin_user_success(self):
        """User with can_add_typification permission can add a new pattern."""
        data = {"sentence": "NEW PATTERN"}
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify pattern was created
        new_pattern = Pattern.objects.filter(sentence="NEW PATTERN").first()
        self.assertIsNotNone(new_pattern)
        self.assertTrue(new_pattern.is_active)
        self.assertEqual(new_pattern.typification, self.target_typification)

    def test_post_add_admin_user_failure(self):
        """User with can_add_typification permission gets form errors on invalid data."""
        data = {"sentence": ""}  # Empty sentence should fail validation
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify pattern was not created
        self.assertFalse(Pattern.objects.filter(sentence="").exists())

    def test_post_add_list_only_user_forbidden(self):
        """User with only list permission cannot add a new pattern."""
        data = {"sentence": "SHOULD NOT BE CREATED"}
        response = self.list_only_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify pattern was not created
        self.assertFalse(Pattern.objects.filter(sentence="SHOULD NOT BE CREATED").exists())

    def test_post_add_no_perm_user_forbidden(self):
        """User with no permissions cannot add a new pattern."""
        data = {"sentence": "SHOULD NOT BE CREATED"}
        response = self.no_perm_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify pattern was not created
        self.assertFalse(Pattern.objects.filter(sentence="SHOULD NOT BE CREATED").exists())

    def test_post_add_anonymous_user_redirect(self):
        """Test that an anonymous user cannot add a new pattern."""
        data = {"sentence": "SHOULD NOT BE CREATED"}
        response = self.anonymous_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify pattern was not created
        self.assertFalse(Pattern.objects.filter(sentence="SHOULD NOT BE CREATED").exists())

    def test_delete_add_bad_request(self):
        """Test that DELETE method is not allowed for add."""
        response = self.admin_client.delete(self.add_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class PatternRelatedEditViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Pattern related views with API_ACTION_EDIT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.edit_url = reverse(
            "analytics:typification:pattern:edit",
            args=[self.target_typification.pk, self.target_pattern.pk],
        )

    def test_get_edit_admin_user(self):
        """User with can_edit_typification permission can access the edit form."""
        response = self.admin_client.get(self.edit_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, self.target_pattern.sentence)

    def test_get_edit_list_only_user(self):
        """User with only list permission cannot access the edit form."""
        response = self.list_only_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_no_perm_user(self):
        """User with no permissions cannot access the edit form."""
        response = self.no_perm_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_anonymous_user(self):
        """Test that an anonymous user cannot access the edit form."""
        response = self.anonymous_client.get(self.edit_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_edit_admin_user_success(self):
        """User with can_edit_typification permission can edit a pattern."""
        data = {"sentence": "UPDATED PATTERN"}
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify pattern was updated
        self.target_pattern.refresh_from_db()
        self.assertEqual(self.target_pattern.sentence, "UPDATED PATTERN")

    def test_post_edit_admin_user_failure(self):
        """User with can_edit_typification permission gets form errors on invalid data."""
        data = {"sentence": ""}  # Empty sentence should fail validation
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify pattern was not updated
        self.target_pattern.refresh_from_db()
        self.assertEqual(self.target_pattern.sentence, "TARGET PATTERN")

    def test_post_edit_list_only_user_forbidden(self):
        """User with only list permission cannot edit a pattern."""
        data = {"sentence": "SHOULD NOT BE UPDATED"}
        response = self.list_only_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify pattern was not updated
        self.target_pattern.refresh_from_db()
        self.assertEqual(self.target_pattern.sentence, "TARGET PATTERN")

    def test_post_edit_no_perm_user_forbidden(self):
        """User with no permissions cannot edit a pattern."""
        data = {"sentence": "SHOULD NOT BE UPDATED"}
        response = self.no_perm_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify pattern was not updated
        self.target_pattern.refresh_from_db()
        self.assertEqual(self.target_pattern.sentence, "TARGET PATTERN")

    def test_post_edit_anonymous_user_redirect(self):
        """Test that an anonymous user cannot edit a pattern."""
        data = {"sentence": "SHOULD NOT BE UPDATED"}
        response = self.anonymous_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify pattern was not updated
        self.target_pattern.refresh_from_db()
        self.assertEqual(self.target_pattern.sentence, "TARGET PATTERN")

    def test_delete_edit_bad_request(self):
        """Test that DELETE method is not allowed for edit."""
        response = self.admin_client.delete(self.edit_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class PatternRelatedDeleteViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Pattern related views with API_ACTION_DELETE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.delete_url = reverse(
            "analytics:typification:pattern:delete",
            args=[self.target_typification.pk, self.target_pattern.pk],
        )

    def test_delete_admin_user_success(self):
        """User with can_delete_typification permission can delete a pattern."""
        response = self.admin_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify pattern was soft-deleted
        self.target_pattern.refresh_from_db()
        self.assertFalse(self.target_pattern.is_active)

    def test_delete_list_only_user_forbidden(self):
        """User with only list permission cannot delete a pattern."""
        response = self.list_only_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify pattern was not deleted
        self.target_pattern.refresh_from_db()
        self.assertTrue(self.target_pattern.is_active)

    def test_delete_no_perm_user_forbidden(self):
        """User with no permissions cannot delete a pattern."""
        response = self.no_perm_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify pattern was not deleted
        self.target_pattern.refresh_from_db()
        self.assertTrue(self.target_pattern.is_active)

    def test_delete_anonymous_user_redirect(self):
        """Test that an anonymous user cannot delete a pattern."""
        response = self.anonymous_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify pattern was not deleted
        self.target_pattern.refresh_from_db()
        self.assertTrue(self.target_pattern.is_active)

    def test_get_delete_bad_request(self):
        """Test that GET method is not allowed for delete."""
        response = self.admin_client.get(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_post_delete_bad_request(self):
        """Test that POST method is not allowed for delete."""
        response = self.admin_client.post(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class PatternRelatedReactivateViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Pattern related views with API_ACTION_REACTIVATE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.reactivate_url = reverse(
            "analytics:typification:pattern:reactivate",
            args=[self.target_typification.pk, self.deleted_pattern.pk],
        )

    def test_post_reactivate_admin_user_success(self):
        """User with can_delete_typification permission can reactivate a pattern."""
        response = self.admin_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify pattern was reactivated
        self.deleted_pattern.refresh_from_db()
        self.assertTrue(self.deleted_pattern.is_active)

    def test_post_reactivate_list_only_user_forbidden(self):
        """User with only list permission cannot reactivate a pattern."""
        response = self.list_only_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify pattern was not reactivated
        self.deleted_pattern.refresh_from_db()
        self.assertFalse(self.deleted_pattern.is_active)

    def test_post_reactivate_no_perm_user_forbidden(self):
        """User with no permissions cannot reactivate a pattern."""
        response = self.no_perm_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify pattern was not reactivated
        self.deleted_pattern.refresh_from_db()
        self.assertFalse(self.deleted_pattern.is_active)

    def test_post_reactivate_anonymous_user_redirect(self):
        """Test that an anonymous user cannot reactivate a pattern."""
        response = self.anonymous_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify pattern was not reactivated
        self.deleted_pattern.refresh_from_db()
        self.assertFalse(self.deleted_pattern.is_active)

    def test_get_reactivate_bad_request(self):
        """Test that GET method is not allowed for reactivate."""
        response = self.admin_client.get(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_reactivate_bad_request(self):
        """Test that DELETE method is not allowed for reactivate."""
        response = self.admin_client.delete(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class PatternRelatedReadViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Pattern related views with API_ACTION_READ action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.read_url = reverse(
            "analytics:typification:pattern:read",
            args=[self.target_typification.pk, self.target_pattern.pk],
        )

    def test_get_read_admin_user(self):
        """User with can_list_typification permission can access the read view."""
        response = self.admin_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_pattern.sentence)
        # Check that the form is read-only (no submit button)
        self.assertContains(response, "disabled")

    def test_get_read_list_only_user(self):
        """User with only list permission can access the read view."""
        response = self.list_only_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_pattern.sentence)
        # Check that the form is read-only (no submit button)
        self.assertContains(response, "disabled")

    def test_get_read_no_perm_user_forbidden(self):
        """User with no permissions cannot access the read view."""
        response = self.no_perm_client.get(self.read_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_read_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the read view."""
        response = self.anonymous_client.get(self.read_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_read_bad_request(self):
        """Test that POST method is not allowed for read."""
        response = self.admin_client.post(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_read_bad_request(self):
        """Test that DELETE method is not allowed for read."""
        response = self.admin_client.delete(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class PatternRelatedHistoryViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Pattern related views with API_ACTION_HISTORY action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.history_url = reverse(
            "analytics:typification:pattern:history",
            args=[self.target_typification.pk, self.target_pattern.pk],
        )

    def test_get_history_admin_user(self):
        """User with can_history_typification permission can access the history view."""
        response = self.admin_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion component

    def test_get_history_list_only_user(self):
        """User with only list permission can access the history view."""
        response = self.list_only_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion component

    def test_get_history_no_perm_user_forbidden(self):
        """User with no permissions cannot access the history view."""
        response = self.no_perm_client.get(self.history_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_history_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the history view."""
        response = self.anonymous_client.get(self.history_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_history_bad_request(self):
        """Test that POST method is not allowed for history."""
        response = self.admin_client.post(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_history_bad_request(self):
        """Test that DELETE method is not allowed for history."""
        response = self.admin_client.delete(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class PatternRelatedListViewTestCase(TypificationViewsBaseTestCase):
    """Test case for the Pattern related views with API_ACTION_LIST action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.list_url = reverse(
            "analytics:typification:pattern:list", args=[self.target_typification.pk]
        )

    def test_get_list_admin_user(self):
        """User with can_list_typification permission can access the list view."""
        response = self.admin_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the table contains the target pattern
        self.assertContains(response, self.target_pattern.sentence)
        # Check that the table contains the deleted pattern
        self.assertContains(response, self.deleted_pattern.sentence)

    def test_get_list_list_only_user(self):
        """User with only list permission can access the list view."""
        response = self.list_only_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the table contains the target pattern
        self.assertContains(response, self.target_pattern.sentence)
        # Check that the table contains the deleted pattern
        self.assertContains(response, self.deleted_pattern.sentence)

    def test_get_list_no_perm_user_forbidden(self):
        """User with no permissions cannot access the list view."""
        response = self.no_perm_client.get(self.list_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_list_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the list view."""
        response = self.anonymous_client.get(self.list_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_list_bad_request(self):
        """Test that POST method is not allowed for list."""
        response = self.admin_client.post(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_list_bad_request(self):
        """Test that DELETE method is not allowed for list."""
        response = self.admin_client.delete(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method
