import io

from django.contrib.auth.hashers import make_password
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.urls import reverse

import openpyxl

from apps.analytics.models import Word, WordList
from apps.common.constants import SCOPE_CAMPAIGN, SCOPE_NONE
from apps.users.models import Campaign, Role, User


class WordListViewsBaseTestCase(TestCase):
    """Base test case for WordList views with various API actions."""

    fixtures = ["prod/roles.json", "prod/campaigns.json", "prod/users.json"]

    def setUp(self):
        """Set up test data."""
        # Create roles with different permissions
        self.admin_role = Role.objects.create(
            name="TEST ADMIN ROLE",
            is_active=True,
            can_add_wordlist=True,
            can_edit_wordlist=True,
            can_delete_wordlist=True,
            can_export_wordlist=True,
            can_history_wordlist=True,
            can_import_wordlist=True,
            scope_wordlist=SCOPE_CAMPAIGN,
        )

        self.list_only_role = Role.objects.create(
            name="TEST LIST ONLY ROLE",
            is_active=True,
            can_add_wordlist=False,
            can_edit_wordlist=False,
            can_delete_wordlist=False,
            can_export_wordlist=False,
            can_history_wordlist=True,
            scope_wordlist=SCOPE_CAMPAIGN,
        )

        self.no_perm_role = Role.objects.create(
            name="TEST NO PERM ROLE",
            is_active=True,
            can_add_wordlist=False,
            can_edit_wordlist=False,
            can_delete_wordlist=False,
            can_export_wordlist=False,
            can_history_wordlist=False,
            scope_wordlist=SCOPE_NONE,
        )

        # Create a test campaign
        self.test_campaign = Campaign.objects.create(
            name="TEST CAMPAIGN", description="Campaign for testing", is_active=True
        )

        # Create users with different roles
        self.admin_user = User.objects.create(
            username="99999999",
            password=make_password("password"),
            first_name="ADMIN",
            last_name="USER",
            document_type=1,
            document_number="99999999",
            role=self.admin_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.list_only_user = User.objects.create(
            username="88888888",
            password=make_password("password"),
            first_name="LIST ONLY",
            last_name="USER",
            document_type=1,
            document_number="88888888",
            role=self.list_only_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.no_perm_user = User.objects.create(
            username="77777777",
            password=make_password("password"),
            first_name="NO PERM",
            last_name="USER",
            document_type=1,
            document_number="77777777",
            role=self.no_perm_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a test wordlist that will be the target of actions
        self.target_wordlist = WordList.objects.create(
            name="TARGET WORDLIST",
            create_user=self.admin_user,
            modify_user=self.admin_user,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a test word related to the target wordlist
        self.target_word = Word.objects.create(
            word="TARGET WORD",
            wordlist=self.target_wordlist,
            create_user=self.admin_user,
            modify_user=self.admin_user,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a soft-deleted wordlist for reactivation tests
        self.deleted_wordlist = WordList.objects.create(
            name="DELETED WORDLIST",
            create_user=self.admin_user,
            modify_user=self.admin_user,
            campaign=self.test_campaign,
            is_active=False,
        )

        # Create a soft-deleted word for reactivation tests
        self.deleted_word = Word.objects.create(
            word="DELETED WORD",
            wordlist=self.target_wordlist,
            create_user=self.admin_user,
            modify_user=self.admin_user,
            campaign=self.test_campaign,
            is_active=False,
        )

        # Create clients for each user
        self.admin_client = Client()
        self.list_only_client = Client()
        self.no_perm_client = Client()
        self.anonymous_client = Client()

        # Log in the clients
        self.admin_client.login(username="99999999", password="password")
        self.list_only_client.login(username="88888888", password="password")
        self.no_perm_client.login(username="77777777", password="password")


class WordListAddViewTestCase(WordListViewsBaseTestCase):
    """Test case for the WordList views with API_ACTION_ADD action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.add_url = reverse("analytics:wordlist:add")

    def test_get_add_admin_user(self):
        """User with can_add_wordlist permission can access the add form."""
        response = self.admin_client.get(self.add_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")

    def test_get_add_list_only_user(self):
        """User with only list permission cannot access the add form."""
        response = self.list_only_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_no_perm_user(self):
        """User with no permissions cannot access the add form."""
        response = self.no_perm_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_anonymous_user(self):
        """Test that an anonymous user cannot access the add form."""
        response = self.anonymous_client.get(self.add_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_add_admin_user_success(self):
        """User with can_add_wordlist permission can add a new wordlist."""
        data = {"name": "NEW WORDLIST"}
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify wordlist was created
        new_wordlist = WordList.objects.filter(name="NEW WORDLIST").first()
        self.assertIsNotNone(new_wordlist)
        self.assertTrue(new_wordlist.is_active)

    def test_post_add_admin_user_failure(self):
        """User with can_add_wordlist permission gets form errors on invalid data."""
        data = {"name": ""}  # Empty name should fail validation
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify wordlist was not created
        self.assertFalse(WordList.objects.filter(name="").exists())

    def test_post_add_list_only_user_forbidden(self):
        """User with only list permission cannot add a new wordlist."""
        data = {"name": "SHOULD NOT BE CREATED"}
        response = self.list_only_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify wordlist was not created
        self.assertFalse(WordList.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_post_add_no_perm_user_forbidden(self):
        """User with no permissions cannot add a new wordlist."""
        data = {"name": "SHOULD NOT BE CREATED"}
        response = self.no_perm_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify wordlist was not created
        self.assertFalse(WordList.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_post_add_anonymous_user_redirect(self):
        """Test that an anonymous user cannot add a new wordlist."""
        data = {"name": "SHOULD NOT BE CREATED"}
        response = self.anonymous_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify wordlist was not created
        self.assertFalse(WordList.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_delete_add_bad_request(self):
        """Test that DELETE method is not allowed for add."""
        response = self.admin_client.delete(self.add_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordListHomeViewTestCase(WordListViewsBaseTestCase):
    """Test case for the WordList views with API_ACTION_HOME action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.home_url = reverse("analytics:wordlist:home")

    def test_get_home_admin_user(self):
        """User with can_list_wordlist permission can access the home page."""
        response = self.admin_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        # Validate template used
        self.assertTemplateUsed(response, "analytics/wordlist/home.html")

    def test_get_home_list_only_user(self):
        """User with only list permission can access the home page."""
        response = self.list_only_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "analytics/wordlist/home.html")

    def test_get_home_no_perm_user(self):
        """User with no permissions cannot access the home page."""
        response = self.no_perm_client.get(self.home_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_home_anonymous_user(self):
        """Test that an anonymous user cannot access the home page."""
        response = self.anonymous_client.get(self.home_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_home_bad_request(self):
        """Test that POST method is not allowed for home."""
        response = self.admin_client.post(self.home_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordListEditViewTestCase(WordListViewsBaseTestCase):
    """Test case for the WordList views with API_ACTION_EDIT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.edit_url = reverse("analytics:wordlist:edit", args=[self.target_wordlist.pk])

    def test_get_edit_admin_user(self):
        """User with can_edit_wordlist permission can access the edit form."""
        response = self.admin_client.get(self.edit_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, self.target_wordlist.name)

    def test_get_edit_list_only_user(self):
        """User with only list permission cannot access the edit form."""
        response = self.list_only_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_no_perm_user(self):
        """User with no permissions cannot access the edit form."""
        response = self.no_perm_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_anonymous_user(self):
        """Test that an anonymous user cannot access the edit form."""
        response = self.anonymous_client.get(self.edit_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_edit_admin_user_success(self):
        """User with can_edit_wordlist permission can edit a wordlist."""
        data = {"name": "UPDATED WORDLIST"}
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify wordlist was updated
        self.target_wordlist.refresh_from_db()
        self.assertEqual(self.target_wordlist.name, "UPDATED WORDLIST")

    def test_post_edit_admin_user_failure(self):
        """User with can_edit_wordlist permission gets form errors on invalid data."""
        data = {"name": ""}  # Empty name should fail validation
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify wordlist was not updated
        self.target_wordlist.refresh_from_db()
        self.assertEqual(self.target_wordlist.name, "TARGET WORDLIST")

    def test_post_edit_list_only_user_forbidden(self):
        """User with only list permission cannot edit a wordlist."""
        data = {"name": "SHOULD NOT BE UPDATED"}
        response = self.list_only_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify wordlist was not updated
        self.target_wordlist.refresh_from_db()
        self.assertEqual(self.target_wordlist.name, "TARGET WORDLIST")

    def test_post_edit_no_perm_user_forbidden(self):
        """User with no permissions cannot edit a wordlist."""
        data = {"name": "SHOULD NOT BE UPDATED"}
        response = self.no_perm_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify wordlist was not updated
        self.target_wordlist.refresh_from_db()
        self.assertEqual(self.target_wordlist.name, "TARGET WORDLIST")

    def test_post_edit_anonymous_user_redirect(self):
        """Test that an anonymous user cannot edit a wordlist."""
        data = {"name": "SHOULD NOT BE UPDATED"}
        response = self.anonymous_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify wordlist was not updated
        self.target_wordlist.refresh_from_db()
        self.assertEqual(self.target_wordlist.name, "TARGET WORDLIST")

    def test_delete_edit_bad_request(self):
        """Test that DELETE method is not allowed for edit."""
        response = self.admin_client.delete(self.edit_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordListDeleteViewTestCase(WordListViewsBaseTestCase):
    """Test case for the WordList views with API_ACTION_DELETE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.delete_url = reverse("analytics:wordlist:delete", args=[self.target_wordlist.pk])

    def test_delete_admin_user_success(self):
        """User with can_delete_wordlist permission can delete a wordlist."""
        response = self.admin_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify wordlist was soft-deleted
        self.target_wordlist.refresh_from_db()
        self.assertFalse(self.target_wordlist.is_active)

        # Verify related words were also soft-deleted
        self.target_word.refresh_from_db()
        self.assertFalse(self.target_word.is_active)

    def test_delete_list_only_user_forbidden(self):
        """User with only list permission cannot delete a wordlist."""
        response = self.list_only_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify wordlist was not deleted
        self.target_wordlist.refresh_from_db()
        self.assertTrue(self.target_wordlist.is_active)

        # Verify related words were not deleted
        self.target_word.refresh_from_db()
        self.assertTrue(self.target_word.is_active)

    def test_delete_no_perm_user_forbidden(self):
        """User with no permissions cannot delete a wordlist."""
        response = self.no_perm_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify wordlist was not deleted
        self.target_wordlist.refresh_from_db()
        self.assertTrue(self.target_wordlist.is_active)

        # Verify related words were not deleted
        self.target_word.refresh_from_db()
        self.assertTrue(self.target_word.is_active)

    def test_delete_anonymous_user_redirect(self):
        """Test that an anonymous user cannot delete a wordlist."""
        response = self.anonymous_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify wordlist was not deleted
        self.target_wordlist.refresh_from_db()
        self.assertTrue(self.target_wordlist.is_active)

        # Verify related words were not deleted
        self.target_word.refresh_from_db()
        self.assertTrue(self.target_word.is_active)

    def test_get_delete_bad_request(self):
        """Test that GET method is not allowed for delete."""
        response = self.admin_client.get(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_post_delete_bad_request(self):
        """Test that POST method is not allowed for delete."""
        response = self.admin_client.post(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordListReactivateViewTestCase(WordListViewsBaseTestCase):
    """Test case for the WordList views with API_ACTION_REACTIVATE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.reactivate_url = reverse(
            "analytics:wordlist:reactivate", args=[self.deleted_wordlist.pk]
        )

    def test_post_reactivate_admin_user_success(self):
        """User with can_delete_wordlist permission can reactivate a wordlist."""
        response = self.admin_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify wordlist was reactivated
        self.deleted_wordlist.refresh_from_db()
        self.assertTrue(self.deleted_wordlist.is_active)

    def test_post_reactivate_list_only_user_forbidden(self):
        """User with only list permission cannot reactivate a wordlist."""
        response = self.list_only_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify wordlist was not reactivated
        self.deleted_wordlist.refresh_from_db()
        self.assertFalse(self.deleted_wordlist.is_active)

    def test_post_reactivate_no_perm_user_forbidden(self):
        """User with no permissions cannot reactivate a wordlist."""
        response = self.no_perm_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify wordlist was not reactivated
        self.deleted_wordlist.refresh_from_db()
        self.assertFalse(self.deleted_wordlist.is_active)

    def test_post_reactivate_anonymous_user_redirect(self):
        """Test that an anonymous user cannot reactivate a wordlist."""
        response = self.anonymous_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify wordlist was not reactivated
        self.deleted_wordlist.refresh_from_db()
        self.assertFalse(self.deleted_wordlist.is_active)

    def test_get_reactivate_bad_request(self):
        """Test that GET method is not allowed for reactivate."""
        response = self.admin_client.get(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_reactivate_bad_request(self):
        """Test that DELETE method is not allowed for reactivate."""
        response = self.admin_client.delete(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordListReadViewTestCase(WordListViewsBaseTestCase):
    """Test case for the WordList views with API_ACTION_READ action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.read_url = reverse("analytics:wordlist:read", args=[self.target_wordlist.pk])

    def test_get_read_admin_user(self):
        """User with can_list_wordlist permission can access the read view."""
        response = self.admin_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_wordlist.name)
        # Check that the form is read-only (no submit button)
        self.assertContains(response, "disabled")

    def test_get_read_list_only_user(self):
        """User with only list permission can access the read view."""
        response = self.list_only_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_wordlist.name)
        # Check that the form is read-only (no submit button)
        self.assertContains(response, "disabled")

    def test_get_read_no_perm_user_forbidden(self):
        """User with no permissions cannot access the read view."""
        response = self.no_perm_client.get(self.read_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_read_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the read view."""
        response = self.anonymous_client.get(self.read_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_read_bad_request(self):
        """Test that POST method is not allowed for read."""
        response = self.admin_client.post(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_read_bad_request(self):
        """Test that DELETE method is not allowed for read."""
        response = self.admin_client.delete(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordListHistoryViewTestCase(WordListViewsBaseTestCase):
    """Test case for the WordList views with API_ACTION_HISTORY action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.history_url = reverse("analytics:wordlist:history", args=[self.target_wordlist.pk])

    def test_get_history_admin_user(self):
        """User with can_history_wordlist permission can access the history view."""
        response = self.admin_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion component

    def test_get_history_list_only_user(self):
        """User with only list permission can access the history view."""
        response = self.list_only_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion component

    def test_get_history_no_perm_user_forbidden(self):
        """User with no permissions cannot access the history view."""
        response = self.no_perm_client.get(self.history_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_history_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the history view."""
        response = self.anonymous_client.get(self.history_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_history_bad_request(self):
        """Test that POST method is not allowed for history."""
        response = self.admin_client.post(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_history_bad_request(self):
        """Test that DELETE method is not allowed for history."""
        response = self.admin_client.delete(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordListListViewTestCase(WordListViewsBaseTestCase):
    """Test case for the WordList views with API_ACTION_LIST action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.list_url = reverse("analytics:wordlist:list")

    def test_get_list_admin_user(self):
        """User with can_list_wordlist permission can access the list view."""
        response = self.admin_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the table contains the target wordlist
        self.assertContains(response, self.target_wordlist.name)
        # Check that the table contains the deleted wordlist
        self.assertContains(response, self.deleted_wordlist.name)

    def test_get_list_list_only_user(self):
        """User with only list permission can access the list view."""
        response = self.list_only_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the table contains the target wordlist
        self.assertContains(response, self.target_wordlist.name)
        # Check that the table contains the deleted wordlist
        self.assertContains(response, self.deleted_wordlist.name)

    def test_get_list_no_perm_user_forbidden(self):
        """User with no permissions cannot access the list view."""
        response = self.no_perm_client.get(self.list_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_list_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the list view."""
        response = self.anonymous_client.get(self.list_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_list_bad_request(self):
        """Test that POST method is not allowed for list."""
        response = self.admin_client.post(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_list_bad_request(self):
        """Test that DELETE method is not allowed for list."""
        response = self.admin_client.delete(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordListExportIndividualViewTestCase(WordListViewsBaseTestCase):
    """Test case for the WordList views with API_ACTION_EXPORT_INDIVIDUAL action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.export_individual_url = reverse(
            "analytics:wordlist:export_individual", args=[self.target_wordlist.pk]
        )

    def test_get_export_individual_admin_user(self):
        """User with can_export_wordlist permission can export a wordlist."""
        response = self.admin_client.get(self.export_individual_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(response.has_header("Content-Disposition"))

        # Verify the Excel file contains the expected data
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = wb.active

        # Check that the header row contains "Palabra" column
        header_row = [cell.value for cell in sheet[1]]
        self.assertIn("Palabra", header_row)

        # Check that the data rows contain the word
        word_column_index = header_row.index("Palabra") + 1  # +1 because openpyxl is 1-indexed
        words = [
            sheet.cell(row=i, column=word_column_index).value for i in range(2, sheet.max_row + 1)
        ]
        self.assertIn(self.target_word.word, words)

    def test_get_export_individual_list_only_user_forbidden(self):
        """User with only list permission cannot export a wordlist."""
        response = self.list_only_client.get(self.export_individual_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_individual_no_perm_user_forbidden(self):
        """User with no permissions cannot export a wordlist."""
        response = self.no_perm_client.get(self.export_individual_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_individual_anonymous_user_redirect(self):
        """Test that an anonymous user cannot export a wordlist."""
        response = self.anonymous_client.get(self.export_individual_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_export_individual_bad_request(self):
        """Test that POST method is not allowed for export_individual."""
        response = self.admin_client.post(self.export_individual_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_export_individual_bad_request(self):
        """Test that DELETE method is not allowed for export_individual."""
        response = self.admin_client.delete(self.export_individual_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordListImportViewTestCase(WordListViewsBaseTestCase):
    """Test case for the WordList views with API_ACTION_IMPORT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.import_url = reverse("analytics:wordlist:import")

    def test_get_import_admin_user(self):
        """User with can_import_wordlist permission can access the import form."""
        response = self.admin_client.get(self.import_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, "file")  # Check for file upload field

    def test_get_import_list_only_user(self):
        """User with only list permission cannot access the import form."""
        response = self.list_only_client.get(self.import_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_import_no_perm_user(self):
        """User with no permissions cannot access the import form."""
        response = self.no_perm_client.get(self.import_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_import_anonymous_user(self):
        """Test that an anonymous user cannot access the import form."""
        response = self.anonymous_client.get(self.import_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_import_admin_user_success(self):
        """User with can_import_wordlist permission can import a wordlist."""
        # Create a test Excel file
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["word"])  # Header row
        sheet.append(["IMPORTED WORD 1"])
        sheet.append(["IMPORTED WORD 2"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Submit the form with the Excel file
        data = {
            "name": "IMPORTED WORDLIST",
            "file": SimpleUploadedFile(
                "test.xlsx",
                excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = self.admin_client.post(self.import_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify wordlist was created
        imported_wordlist = WordList.objects.filter(name="IMPORTED WORDLIST").first()
        self.assertIsNotNone(imported_wordlist)
        self.assertTrue(imported_wordlist.is_active)

        # Verify words were created
        imported_words = Word.objects.filter(wordlist=imported_wordlist)
        self.assertEqual(imported_words.count(), 2)
        self.assertIn("IMPORTED WORD 1", [word.word for word in imported_words])
        self.assertIn("IMPORTED WORD 2", [word.word for word in imported_words])

    def test_post_import_list_only_user_forbidden(self):
        """User with only list permission cannot import a wordlist."""
        # Create a test Excel file

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["word"])  # Header row
        sheet.append(["SHOULD NOT BE IMPORTED"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Submit the form with the Excel file
        data = {
            "name": "SHOULD NOT BE IMPORTED",
            "file": SimpleUploadedFile(
                "test.xlsx",
                excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = self.list_only_client.post(self.import_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify wordlist was not created
        self.assertFalse(WordList.objects.filter(name="SHOULD NOT BE IMPORTED").exists())

    def test_post_import_no_perm_user_forbidden(self):
        """User with no permissions cannot import a wordlist."""
        # Create a test Excel file
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["word"])  # Header row
        sheet.append(["SHOULD NOT BE IMPORTED"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Submit the form with the Excel file
        data = {
            "name": "SHOULD NOT BE IMPORTED",
            "file": SimpleUploadedFile(
                "test.xlsx",
                excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = self.no_perm_client.post(self.import_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify wordlist was not created
        self.assertFalse(WordList.objects.filter(name="SHOULD NOT BE IMPORTED").exists())

    def test_post_import_anonymous_user_redirect(self):
        """Test that an anonymous user cannot import a wordlist."""
        # Create a test Excel file
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["word"])  # Header row
        sheet.append(["SHOULD NOT BE IMPORTED"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Submit the form with the Excel file
        data = {
            "name": "SHOULD NOT BE IMPORTED",
            "file": SimpleUploadedFile(
                "test.xlsx",
                excel_file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        response = self.anonymous_client.post(self.import_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify wordlist was not created
        self.assertFalse(WordList.objects.filter(name="SHOULD NOT BE IMPORTED").exists())

    def test_delete_import_bad_request(self):
        """Test that DELETE method is not allowed for import."""
        response = self.admin_client.delete(self.import_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordRelatedAddViewTestCase(WordListViewsBaseTestCase):
    """Test case for the Word related views with API_ACTION_ADD action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.add_url = reverse("analytics:wordlist:word:add", args=[self.target_wordlist.pk])

    def test_get_add_admin_user(self):
        """User with can_add_wordlist permission can access the add form."""
        response = self.admin_client.get(self.add_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")

    def test_get_add_list_only_user(self):
        """User with only list permission cannot access the add form."""
        response = self.list_only_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_no_perm_user(self):
        """User with no permissions cannot access the add form."""
        response = self.no_perm_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_anonymous_user(self):
        """Test that an anonymous user cannot access the add form."""
        response = self.anonymous_client.get(self.add_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_add_admin_user_success(self):
        """User with can_add_wordlist permission can add a new word."""
        data = {"word": "NEW WORD"}
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify word was created
        new_word = Word.objects.filter(word="NEW WORD").first()
        self.assertIsNotNone(new_word)
        self.assertTrue(new_word.is_active)
        self.assertEqual(new_word.wordlist, self.target_wordlist)

    def test_post_add_admin_user_failure(self):
        """User with can_add_wordlist permission gets form errors on invalid data."""
        data = {"word": ""}  # Empty word should fail validation
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify word was not created
        self.assertFalse(Word.objects.filter(word="").exists())

    def test_post_add_list_only_user_forbidden(self):
        """User with only list permission cannot add a new word."""
        data = {"word": "SHOULD NOT BE CREATED"}
        response = self.list_only_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify word was not created
        self.assertFalse(Word.objects.filter(word="SHOULD NOT BE CREATED").exists())

    def test_post_add_no_perm_user_forbidden(self):
        """User with no permissions cannot add a new word."""
        data = {"word": "SHOULD NOT BE CREATED"}
        response = self.no_perm_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify word was not created
        self.assertFalse(Word.objects.filter(word="SHOULD NOT BE CREATED").exists())

    def test_post_add_anonymous_user_redirect(self):
        """Test that an anonymous user cannot add a new word."""
        data = {"word": "SHOULD NOT BE CREATED"}
        response = self.anonymous_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify word was not created
        self.assertFalse(Word.objects.filter(word="SHOULD NOT BE CREATED").exists())

    def test_delete_add_bad_request(self):
        """Test that DELETE method is not allowed for add."""
        response = self.admin_client.delete(self.add_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordRelatedEditViewTestCase(WordListViewsBaseTestCase):
    """Test case for the Word related views with API_ACTION_EDIT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.edit_url = reverse(
            "analytics:wordlist:word:edit", args=[self.target_wordlist.pk, self.target_word.pk]
        )

    def test_get_edit_admin_user(self):
        """User with can_edit_wordlist permission can access the edit form."""
        response = self.admin_client.get(self.edit_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, self.target_word.word)

    def test_get_edit_list_only_user(self):
        """User with only list permission cannot access the edit form."""
        response = self.list_only_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_no_perm_user(self):
        """User with no permissions cannot access the edit form."""
        response = self.no_perm_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_anonymous_user(self):
        """Test that an anonymous user cannot access the edit form."""
        response = self.anonymous_client.get(self.edit_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_edit_admin_user_success(self):
        """User with can_edit_wordlist permission can edit a word."""
        data = {"word": "UPDATED WORD"}
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify word was updated
        self.target_word.refresh_from_db()
        self.assertEqual(self.target_word.word, "UPDATED WORD")

    def test_post_edit_admin_user_failure(self):
        """User with can_edit_wordlist permission gets form errors on invalid data."""
        data = {"word": ""}  # Empty word should fail validation
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify word was not updated
        self.target_word.refresh_from_db()
        self.assertEqual(self.target_word.word, "TARGET WORD")

    def test_post_edit_list_only_user_forbidden(self):
        """User with only list permission cannot edit a word."""
        data = {"word": "SHOULD NOT BE UPDATED"}
        response = self.list_only_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify word was not updated
        self.target_word.refresh_from_db()
        self.assertEqual(self.target_word.word, "TARGET WORD")

    def test_post_edit_no_perm_user_forbidden(self):
        """User with no permissions cannot edit a word."""
        data = {"word": "SHOULD NOT BE UPDATED"}
        response = self.no_perm_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify word was not updated
        self.target_word.refresh_from_db()
        self.assertEqual(self.target_word.word, "TARGET WORD")

    def test_post_edit_anonymous_user_redirect(self):
        """Test that an anonymous user cannot edit a word."""
        data = {"word": "SHOULD NOT BE UPDATED"}
        response = self.anonymous_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify word was not updated
        self.target_word.refresh_from_db()
        self.assertEqual(self.target_word.word, "TARGET WORD")

    def test_delete_edit_bad_request(self):
        """Test that DELETE method is not allowed for edit."""
        response = self.admin_client.delete(self.edit_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordRelatedDeleteViewTestCase(WordListViewsBaseTestCase):
    """Test case for the Word related views with API_ACTION_DELETE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.delete_url = reverse(
            "analytics:wordlist:word:delete", args=[self.target_wordlist.pk, self.target_word.pk]
        )

    def test_delete_admin_user_success(self):
        """User with can_delete_wordlist permission can delete a word."""
        response = self.admin_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify word was soft-deleted
        self.target_word.refresh_from_db()
        self.assertFalse(self.target_word.is_active)

    def test_delete_list_only_user_forbidden(self):
        """User with only list permission cannot delete a word."""
        response = self.list_only_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify word was not deleted
        self.target_word.refresh_from_db()
        self.assertTrue(self.target_word.is_active)

    def test_delete_no_perm_user_forbidden(self):
        """User with no permissions cannot delete a word."""
        response = self.no_perm_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify word was not deleted
        self.target_word.refresh_from_db()
        self.assertTrue(self.target_word.is_active)

    def test_delete_anonymous_user_redirect(self):
        """Test that an anonymous user cannot delete a word."""
        response = self.anonymous_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify word was not deleted
        self.target_word.refresh_from_db()
        self.assertTrue(self.target_word.is_active)

    def test_get_delete_bad_request(self):
        """Test that GET method is not allowed for delete."""
        response = self.admin_client.get(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_post_delete_bad_request(self):
        """Test that POST method is not allowed for delete."""
        response = self.admin_client.post(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordRelatedReactivateViewTestCase(WordListViewsBaseTestCase):
    """Test case for the Word related views with API_ACTION_REACTIVATE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.reactivate_url = reverse(
            "analytics:wordlist:word:reactivate",
            args=[self.target_wordlist.pk, self.deleted_word.pk],
        )

    def test_post_reactivate_admin_user_success(self):
        """User with can_delete_wordlist permission can reactivate a word."""
        response = self.admin_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify word was reactivated
        self.deleted_word.refresh_from_db()
        self.assertTrue(self.deleted_word.is_active)

    def test_post_reactivate_list_only_user_forbidden(self):
        """User with only list permission cannot reactivate a word."""
        response = self.list_only_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify word was not reactivated
        self.deleted_word.refresh_from_db()
        self.assertFalse(self.deleted_word.is_active)

    def test_post_reactivate_no_perm_user_forbidden(self):
        """User with no permissions cannot reactivate a word."""
        response = self.no_perm_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify word was not reactivated
        self.deleted_word.refresh_from_db()
        self.assertFalse(self.deleted_word.is_active)

    def test_post_reactivate_anonymous_user_redirect(self):
        """Test that an anonymous user cannot reactivate a word."""
        response = self.anonymous_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify word was not reactivated
        self.deleted_word.refresh_from_db()
        self.assertFalse(self.deleted_word.is_active)

    def test_get_reactivate_bad_request(self):
        """Test that GET method is not allowed for reactivate."""
        response = self.admin_client.get(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_reactivate_bad_request(self):
        """Test that DELETE method is not allowed for reactivate."""
        response = self.admin_client.delete(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordRelatedReadViewTestCase(WordListViewsBaseTestCase):
    """Test case for the Word related views with API_ACTION_READ action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.read_url = reverse(
            "analytics:wordlist:word:read", args=[self.target_wordlist.pk, self.target_word.pk]
        )

    def test_get_read_admin_user(self):
        """User with can_list_wordlist permission can access the read view."""
        response = self.admin_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_word.word)
        # Check that the form is read-only (no submit button)
        self.assertContains(response, "disabled")

    def test_get_read_list_only_user(self):
        """User with only list permission can access the read view."""
        response = self.list_only_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_word.word)
        # Check that the form is read-only (no submit button)
        self.assertContains(response, "disabled")

    def test_get_read_no_perm_user_forbidden(self):
        """User with no permissions cannot access the read view."""
        response = self.no_perm_client.get(self.read_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_read_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the read view."""
        response = self.anonymous_client.get(self.read_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_read_bad_request(self):
        """Test that POST method is not allowed for read."""
        response = self.admin_client.post(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_read_bad_request(self):
        """Test that DELETE method is not allowed for read."""
        response = self.admin_client.delete(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordRelatedHistoryViewTestCase(WordListViewsBaseTestCase):
    """Test case for the Word related views with API_ACTION_HISTORY action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.history_url = reverse(
            "analytics:wordlist:word:history", args=[self.target_wordlist.pk, self.target_word.pk]
        )

    def test_get_history_admin_user(self):
        """User with can_history_wordlist permission can access the history view."""
        response = self.admin_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion component

    def test_get_history_list_only_user(self):
        """User with only list permission can access the history view."""
        response = self.list_only_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion component

    def test_get_history_no_perm_user_forbidden(self):
        """User with no permissions cannot access the history view."""
        response = self.no_perm_client.get(self.history_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_history_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the history view."""
        response = self.anonymous_client.get(self.history_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_history_bad_request(self):
        """Test that POST method is not allowed for history."""
        response = self.admin_client.post(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_history_bad_request(self):
        """Test that DELETE method is not allowed for history."""
        response = self.admin_client.delete(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class WordRelatedListViewTestCase(WordListViewsBaseTestCase):
    """Test case for the Word related views with API_ACTION_LIST action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.list_url = reverse("analytics:wordlist:word:list", args=[self.target_wordlist.pk])

    def test_get_list_admin_user(self):
        """User with can_list_wordlist permission can access the list view."""
        response = self.admin_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the table contains the target word
        self.assertContains(response, self.target_word.word)
        # Check that the table contains the deleted word
        self.assertContains(response, self.deleted_word.word)

    def test_get_list_list_only_user(self):
        """User with only list permission can access the list view."""
        response = self.list_only_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the table contains the target word
        self.assertContains(response, self.target_word.word)
        # Check that the table contains the deleted word
        self.assertContains(response, self.deleted_word.word)

    def test_get_list_no_perm_user_forbidden(self):
        """User with no permissions cannot access the list view."""
        response = self.no_perm_client.get(self.list_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_list_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the list view."""
        response = self.anonymous_client.get(self.list_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_list_bad_request(self):
        """Test that POST method is not allowed for list."""
        response = self.admin_client.post(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_list_bad_request(self):
        """Test that DELETE method is not allowed for list."""
        response = self.admin_client.delete(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method
