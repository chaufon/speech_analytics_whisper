"""
Tests for the User views with various API actions.
"""

import io

from django.contrib.auth.hashers import make_password
from django.test import Client, TestCase
from django.urls import reverse

import openpyxl

from apps.common.constants import SCOPE_CAMPAIGN, SCOPE_GLOBAL, SCOPE_NONE
from apps.users.models import Campaign, Role, User


class UserViewsBaseTestCase(TestCase):
    """Base test case for User views with various API actions."""

    fixtures = ["prod/roles.json", "prod/campaigns.json", "prod/users.json"]

    def setUp(self):
        """Set up test data."""
        # Create a new campaign for testing
        self.test_campaign = Campaign.objects.create(
            name="TEST CAMPAIGN", description="Campaign for testing", is_active=True
        )

        # Create another campaign for cross-campaign testing
        self.other_campaign = Campaign.objects.create(
            name="OTHER CAMPAIGN", description="Another campaign for testing", is_active=True
        )

        # Create a new role with SCOPE_GLOBAL for testing
        self.global_role = Role.objects.create(
            name="TEST GLOBAL ROLE",
            is_active=True,
            can_add_user=True,
            can_edit_user=True,
            can_delete_user=True,
            can_history_user=True,
            can_change_user_password=True,
            can_export_user=True,
            can_import_user=True,
            scope_user=SCOPE_GLOBAL,
        )

        # Create a new role with SCOPE_CAMPAIGN for testing
        self.campaign_role = Role.objects.create(
            name="TEST CAMPAIGN ROLE",
            is_active=True,
            can_add_user=True,
            can_edit_user=True,
            can_delete_user=True,
            can_history_user=True,
            can_change_user_password=True,
            can_export_user=True,
            can_import_user=True,
            scope_user=SCOPE_CAMPAIGN,
        )

        # Create a new role with SCOPE_NONE for testing
        self.none_role = Role.objects.create(
            name="TEST NONE ROLE",
            is_active=True,
            can_add_user=False,
            can_edit_user=False,
            can_delete_user=False,
            can_history_user=False,
            can_change_user_password=False,
            can_export_user=False,
            can_import_user=False,
            scope_user=SCOPE_NONE,
        )

        # Create a role with only list permission
        # Some users have only perm `can_history_user` enabled, to allow them `can_list_user`
        self.list_only_role = Role.objects.create(
            name="TEST LIST ONLY ROLE",
            is_active=True,
            can_add_user=False,
            can_edit_user=False,
            can_delete_user=False,
            can_history_user=True,
            can_change_user_password=False,
            can_export_user=False,
            can_import_user=False,
            scope_user=SCOPE_GLOBAL,
        )

        # Create users with different roles and scopes
        self.global_user = User.objects.create(
            username="99999999",
            password=make_password("password"),
            first_name="GLOBAL",
            last_name="USER",
            document_type=1,
            document_number="99999999",
            role=self.global_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.campaign_user = User.objects.create(
            username="88888888",
            password=make_password("password"),
            first_name="CAMPAIGN",
            last_name="USER",
            document_type=1,
            document_number="88888888",
            role=self.campaign_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.none_user = User.objects.create(
            username="77777777",
            password=make_password("password"),
            first_name="NONE",
            last_name="USER",
            document_type=1,
            document_number="77777777",
            role=self.none_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.list_only_user = User.objects.create(
            username="66666666",
            password=make_password("password"),
            first_name="LIST",
            last_name="ONLY USER",
            document_type=1,
            document_number="66666666",
            role=self.list_only_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a user in the other campaign
        self.other_campaign_user = User.objects.create(
            username="55555555",
            password=make_password("password"),
            first_name="OTHER",
            last_name="CAMPAIGN USER",
            document_type=1,
            document_number="55555555",
            role=self.campaign_role,
            campaign=self.other_campaign,
            is_active=True,
        )

        # Create a test user that will be the target of actions
        self.test_user = User.objects.create(
            username="44444444",
            password=make_password("password"),
            first_name="TEST",
            last_name="USER",
            document_type=1,
            document_number="44444444",
            role=self.campaign_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a test user in the other campaign
        self.test_other_campaign_user = User.objects.create(
            username="33333333",
            password=make_password("password"),
            first_name="TEST OTHER",
            last_name="CAMPAIGN USER",
            document_type=1,
            document_number="33333333",
            role=self.campaign_role,
            campaign=self.other_campaign,
            is_active=True,
        )

        # Create a soft-deleted user for reactivation tests
        self.deleted_user = User.objects.create(
            username="22222222",
            password=make_password("password"),
            first_name="DELETED",
            last_name="USER",
            document_type=1,
            document_number="22222222",
            role=self.campaign_role,
            campaign=self.test_campaign,
            is_active=False,
        )

        # Create a soft-deleted user in the other campaign
        self.deleted_other_campaign_user = User.objects.create(
            username="11111111",
            password=make_password("password"),
            first_name="DELETED OTHER",
            last_name="CAMPAIGN USER",
            document_type=1,
            document_number="11111111",
            role=self.campaign_role,
            campaign=self.other_campaign,
            is_active=False,
        )

        # Create clients for each user
        self.global_client = Client()
        self.campaign_client = Client()
        self.none_client = Client()
        self.list_only_client = Client()
        self.anonymous_client = Client()

        # Log in the clients
        self.global_client.login(username="99999999", password="password")
        self.campaign_client.login(username="88888888", password="password")
        self.none_client.login(username="77777777", password="password")
        self.list_only_client.login(username="66666666", password="password")


class UserAddViewTestCase(UserViewsBaseTestCase):
    """Test case for the User views with API_ACTION_ADD action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.add_url = reverse("users:user:add")

    def test_get_add_global_user(self):
        """User with SCOPE_GLOBAL can access the add form."""
        response = self.global_client.get(self.add_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")

    def test_get_add_campaign_user(self):
        """User with SCOPE_CAMPAIGN can access the add form."""
        response = self.campaign_client.get(self.add_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")

    def test_get_add_none_user(self):
        """User with SCOPE_NONE cannot access the add form."""
        response = self.none_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_list_only_user(self):
        """User with only list permission cannot access the add form."""
        response = self.list_only_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_anonymous_user(self):
        """Test that an anonymous user cannot access the add form."""
        response = self.anonymous_client.get(self.add_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_add_global_user_success(self):
        """User with SCOPE_GLOBAL can add a new user."""
        data = {
            "username": "12121210",
            "email": "new.user.global@example.com",
            "first_name": "NEW",
            "last_name": "USER GLOBAL",
            "document_type": 1,
            "document_number": "12345678",
            "password": "newpassword",
            "role": self.campaign_role.pk,
            "campaign": self.test_campaign.pk,
        }
        response = self.global_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Verify user was created
        new_user = User.objects.filter(username="12121210").first()
        self.assertIsNotNone(new_user)
        self.assertEqual(new_user.email, "new.user.global@example.com")
        self.assertEqual(new_user.first_name, "NEW")
        self.assertEqual(new_user.last_name, "USER GLOBAL")
        self.assertEqual(new_user.document_number, "12345678")
        self.assertEqual(new_user.role, self.campaign_role)
        self.assertEqual(new_user.campaign, self.test_campaign)
        self.assertTrue(new_user.is_active)

    def test_post_add_campaign_user_success(self):
        """User with SCOPE_CAMPAIGN can add a new user."""
        data = {
            "username": "12121211",
            "email": "new.user.campaign@example.com",
            "first_name": "NEW",
            "last_name": "USER CAMPAIGN",
            "document_type": 1,
            "document_number": "87654321",
            "password": "newpassword",
            "role": self.campaign_role.pk,
            "campaign": self.test_campaign.pk,
        }
        response = self.campaign_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 204)  # No content with event header

        # Verify user was created
        new_user = User.objects.filter(username="12121211").first()
        self.assertIsNotNone(new_user)
        self.assertEqual(new_user.email, "new.user.campaign@example.com")
        self.assertEqual(new_user.first_name, "NEW")
        self.assertEqual(new_user.last_name, "USER CAMPAIGN")
        self.assertEqual(new_user.document_number, "87654321")
        self.assertEqual(new_user.role, self.campaign_role)
        self.assertEqual(new_user.campaign, self.test_campaign)
        self.assertTrue(new_user.is_active)

    def test_post_add_global_user_other_campaign_restricted(self):
        """A global-scope non-superuser cannot add a user to another campaign: the form's
        campaign field is limited to the editor's own campaign, so another campaign is an
        invalid choice and the form is re-rendered with errors."""
        data = {
            "username": "12121212",
            "email": "new.user.other.campaign@example.com",
            "first_name": "NEW",
            "last_name": "USER OTHER CAMPAIGN",
            "document_type": 1,
            "document_number": "13579246",
            "password": "newpassword",
            "role": self.campaign_role.pk,
            "campaign": self.other_campaign.pk,
        }
        response = self.global_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 200)  # Form re-rendered with errors

        # Verify user was not created
        self.assertFalse(User.objects.filter(username="12121212").exists())

    def test_post_add_campaign_user_other_campaign_failure(self):
        """User with SCOPE_CAMPAIGN cannot add a user to another campaign."""
        data = {
            "username": "12121214",
            "email": "should.not.be.created@example.com",
            "first_name": "SHOULD",
            "last_name": "NOT BE CREATED",
            "document_type": 1,
            "document_number": "24681357",
            "password": "newpassword",
            "role": self.campaign_role.pk,
            "campaign": self.other_campaign.pk,
        }
        response = self.campaign_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 200)  # Form error
        self.assertContains(response, "invalid-feedback")

        # Verify user was not created
        self.assertFalse(User.objects.filter(username="12121214").exists())

    def test_post_add_global_user_failure(self):
        """User with SCOPE_GLOBAL gets form errors on invalid data."""
        data = {
            "username": "12121215",
            "email": "invalid.user@example.com",
            "first_name": "INVALID",
            "last_name": "USER",
            "document_type": 1,
            "document_number": "",  # Empty document number should fail validation
            "password": "newpassword",
            "role": self.campaign_role.pk,
            "campaign": self.test_campaign.pk,
        }
        response = self.global_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify user was not created
        self.assertFalse(User.objects.filter(username="12121215").exists())

    def test_post_add_none_user_forbidden(self):
        """User with SCOPE_NONE cannot add a new user."""
        data = {
            "username": "12121216",
            "email": "none.user.add@example.com",
            "first_name": "NONE",
            "last_name": "USER ADD",
            "document_type": 1,
            "document_number": "11223344",
            "password": "newpassword",
            "role": self.campaign_role.pk,
            "campaign": self.test_campaign.pk,
        }
        response = self.none_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify user was not created
        self.assertFalse(User.objects.filter(username="12121216").exists())

    def test_post_add_list_only_user_forbidden(self):
        """User with only list permission cannot add a new user."""
        data = {
            "username": "12121217",
            "email": "list.only.user.add@example.com",
            "first_name": "LIST ONLY",
            "last_name": "USER ADD",
            "document_type": 1,
            "document_number": "44332211",
            "password": "newpassword",
            "role": self.campaign_role.pk,
            "campaign": self.test_campaign.pk,
        }
        response = self.list_only_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify user was not created
        self.assertFalse(User.objects.filter(username="12121217").exists())

    def test_post_add_anonymous_user_redirect(self):
        """Test that an anonymous user cannot add a new user."""
        data = {
            "username": "12121218",
            "email": "anonymous.user.add@example.com",
            "first_name": "ANONYMOUS",
            "last_name": "USER ADD",
            "document_type": 1,
            "document_number": "55667788",
            "password": "newpassword",
            "role": self.campaign_role.pk,
            "campaign": self.test_campaign.pk,
        }
        response = self.anonymous_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify user was not created
        self.assertFalse(User.objects.filter(username="12121218").exists())

    def test_delete_add_bad_request(self):
        """Test that DELETE method is not allowed for add."""
        response = self.global_client.delete(self.add_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class UserHomeViewTestCase(UserViewsBaseTestCase):
    """Test case for the User views with API_ACTION_HOME action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.home_url = reverse("users:user:home")

    def test_get_home_global_user(self):
        """User with SCOPE_GLOBAL can access the home page."""
        response = self.global_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        # Validate template used
        self.assertTemplateUsed(response, "users/user/home.html")

    def test_get_home_campaign_user(self):
        """User with SCOPE_CAMPAIGN can access the home page."""
        response = self.campaign_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "users/user/home.html")

    def test_get_home_list_only_user(self):
        """User with only list permission can access the home page."""
        response = self.list_only_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "users/user/home.html")

    def test_get_home_none_user(self):
        """User with SCOPE_NONE cannot access the home page."""
        response = self.none_client.get(self.home_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_home_anonymous_user(self):
        """Test that an anonymous user cannot access the home page."""
        response = self.anonymous_client.get(self.home_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_home_bad_request(self):
        """Test that POST method is not allowed for home page."""
        response = self.global_client.post(self.home_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class UserEditViewTestCase(UserViewsBaseTestCase):
    """Test case for the User views with API_ACTION_EDIT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.edit_url = reverse("users:user:edit", args=[self.test_user.pk])
        self.edit_other_campaign_url = reverse(
            "users:user:edit", args=[self.test_other_campaign_user.pk]
        )

    def test_get_edit_global_user(self):
        """User with SCOPE_GLOBAL can access the edit form."""
        response = self.global_client.get(self.edit_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, self.test_user.username)

    def test_get_edit_global_user_other_campaign(self):
        """User with SCOPE_GLOBAL can edit a user from another campaign."""
        response = self.global_client.get(self.edit_other_campaign_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, self.test_other_campaign_user.username)

    def test_get_edit_campaign_user_same_campaign(self):
        """User with SCOPE_CAMPAIGN can edit a user from the same campaign."""
        response = self.campaign_client.get(self.edit_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, self.test_user.username)

    def test_get_edit_campaign_user_other_campaign(self):
        """User with SCOPE_CAMPAIGN cannot edit a user from another campaign."""
        response = self.campaign_client.get(self.edit_other_campaign_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_none_user(self):
        """User with SCOPE_NONE cannot access the edit form."""
        response = self.none_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_list_only_user(self):
        """User with only list permission cannot access the edit form."""
        response = self.list_only_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_anonymous_user(self):
        """Test that an anonymous user cannot access the edit form."""
        response = self.anonymous_client.get(self.edit_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_edit_global_user_success(self):
        """User with SCOPE_GLOBAL can edit a user."""
        data = {
            "username": self.test_user.username,  # Username cannot be changed
            "email": "updated.email@example.com",
            "first_name": "UPDATED",
            "last_name": "USER",
            "document_type": 1,
            "document_number": self.test_user.document_number,
            "role": self.campaign_role.pk,
            "campaign": self.test_campaign.pk,
        }
        response = self.global_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.test_user.refresh_from_db()
        self.assertEqual(self.test_user.email, "updated.email@example.com")
        self.assertEqual(self.test_user.first_name, "UPDATED")

    def test_post_edit_campaign_user_success(self):
        """User with SCOPE_CAMPAIGN can edit a user from the same campaign."""
        data = {
            "username": self.test_user.username,  # Username cannot be changed
            "email": "campaign.updated@example.com",
            "first_name": "CAMPAIGN",
            "last_name": "UPDATED",
            "document_type": 1,
            "document_number": self.test_user.document_number,
            "role": self.campaign_role.pk,
            "campaign": self.test_campaign.pk,
        }
        response = self.campaign_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.test_user.refresh_from_db()
        self.assertEqual(self.test_user.email, "campaign.updated@example.com")
        self.assertEqual(self.test_user.first_name, "CAMPAIGN")

    def test_post_edit_campaign_user_other_campaign_forbidden(self):
        """User with SCOPE_CAMPAIGN cannot edit a user from another campaign."""
        data = {
            "username": self.test_other_campaign_user.username,
            "email": "should.not.update@example.com",
            "first_name": "SHOULD",
            "last_name": "NOT UPDATE",
            "document_type": 1,
            "document_number": self.test_other_campaign_user.document_number,
            "role": self.campaign_role.pk,
            "campaign": self.other_campaign.pk,
        }
        response = self.campaign_client.post(self.edit_other_campaign_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.test_other_campaign_user.refresh_from_db()
        self.assertNotEqual(self.test_other_campaign_user.first_name, "SHOULD")

    def test_post_edit_global_user_failure(self):
        """User with SCOPE_GLOBAL gets form errors on invalid data."""
        data = {
            "username": self.test_user.username,
            "email": "valid.email@example.com",
            "first_name": "VALID",
            "last_name": "USER",
            "document_type": 1,
            "document_number": "",  # Empty document number should fail validation
            "role": self.campaign_role.pk,
            "campaign": self.test_campaign.pk,
        }
        response = self.global_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Refresh from database to verify no changes
        self.test_user.refresh_from_db()
        self.assertNotEqual(self.test_user.first_name, "VALID")

    def test_delete_edit_bad_request(self):
        """Test that DELETE method is not allowed for edit."""
        response = self.global_client.delete(self.edit_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class UserDeleteViewTestCase(UserViewsBaseTestCase):
    """Test case for the User views with API_ACTION_DELETE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.delete_url = reverse("users:user:delete", args=[self.test_user.pk])
        self.delete_other_campaign_url = reverse(
            "users:user:delete", args=[self.test_other_campaign_user.pk]
        )

    def test_delete_global_user_success(self):
        """User with SCOPE_GLOBAL can soft-delete a user."""
        response = self.global_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.test_user.refresh_from_db()
        self.assertFalse(self.test_user.is_active)

    def test_delete_global_user_other_campaign_success(self):
        """User with SCOPE_GLOBAL can soft-delete a user from another campaign."""
        response = self.global_client.delete(self.delete_other_campaign_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.test_other_campaign_user.refresh_from_db()
        self.assertFalse(self.test_other_campaign_user.is_active)

    def test_delete_campaign_user_same_campaign_success(self):
        """User with SCOPE_CAMPAIGN can soft-delete a user from the same campaign."""
        response = self.campaign_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.test_user.refresh_from_db()
        self.assertFalse(self.test_user.is_active)

    def test_delete_campaign_user_other_campaign_forbidden(self):
        """User with SCOPE_CAMPAIGN cannot soft-delete a user from another campaign."""
        response = self.campaign_client.delete(self.delete_other_campaign_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.test_other_campaign_user.refresh_from_db()
        self.assertTrue(self.test_other_campaign_user.is_active)

    def test_delete_none_user_forbidden(self):
        """User with SCOPE_NONE cannot soft-delete a user."""
        response = self.none_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.test_user.refresh_from_db()
        self.assertTrue(self.test_user.is_active)

    def test_delete_list_only_user_forbidden(self):
        """User with only list permission cannot soft-delete a user."""
        response = self.list_only_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.test_user.refresh_from_db()
        self.assertTrue(self.test_user.is_active)

    def test_delete_anonymous_user_redirect(self):
        """Test that an anonymous user cannot soft-delete a user."""
        response = self.anonymous_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.test_user.refresh_from_db()
        self.assertTrue(self.test_user.is_active)

    def test_get_delete_bad_request(self):
        """Test that GET method is not allowed for delete."""
        response = self.global_client.get(self.delete_url)
        self.assertEqual(response.status_code, 400)

    def test_post_delete_bad_request(self):
        """Test that POST method is not allowed for delete."""
        response = self.global_client.post(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class UserReactivateViewTestCase(UserViewsBaseTestCase):
    """Test case for the User views with API_ACTION_REACTIVATE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.reactivate_url = reverse("users:user:reactivate", args=[self.deleted_user.pk])
        self.reactivate_other_campaign_url = reverse(
            "users:user:reactivate", args=[self.deleted_other_campaign_user.pk]
        )

    def test_post_reactivate_global_user_success(self):
        """User with SCOPE_GLOBAL can reactivate a soft-deleted user."""
        response = self.global_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.deleted_user.refresh_from_db()
        self.assertTrue(self.deleted_user.is_active)

    def test_post_reactivate_global_user_other_campaign_success(self):
        """User with SCOPE_GLOBAL can reactivate a soft-deleted user from another campaign."""
        response = self.global_client.post(self.reactivate_other_campaign_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.deleted_other_campaign_user.refresh_from_db()
        self.assertTrue(self.deleted_other_campaign_user.is_active)

    def test_post_reactivate_campaign_user_same_campaign_success(self):
        """User with SCOPE_CAMPAIGN can reactivate a soft-deleted user from the same campaign."""
        response = self.campaign_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.deleted_user.refresh_from_db()
        self.assertTrue(self.deleted_user.is_active)

    def test_post_reactivate_campaign_user_other_campaign_forbidden(self):
        """User with SCOPE_CAMPAIGN cannot reactivate a soft-deleted user from another campaign."""
        response = self.campaign_client.post(self.reactivate_other_campaign_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.deleted_other_campaign_user.refresh_from_db()
        self.assertFalse(self.deleted_other_campaign_user.is_active)

    def test_post_reactivate_none_user_forbidden(self):
        """User with SCOPE_NONE cannot reactivate a soft-deleted user."""
        response = self.none_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.deleted_user.refresh_from_db()
        self.assertFalse(self.deleted_user.is_active)

    def test_post_reactivate_list_only_user_forbidden(self):
        """User with only list permission cannot reactivate a soft-deleted user."""
        response = self.list_only_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.deleted_user.refresh_from_db()
        self.assertFalse(self.deleted_user.is_active)

    def test_post_reactivate_anonymous_user_redirect(self):
        """Test that an anonymous user cannot reactivate a soft-deleted user."""
        response = self.anonymous_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.deleted_user.refresh_from_db()
        self.assertFalse(self.deleted_user.is_active)

    def test_get_reactivate_bad_request(self):
        """Test that GET method is not allowed for reactivate."""
        response = self.global_client.get(self.reactivate_url)
        self.assertEqual(response.status_code, 400)

    def test_delete_reactivate_bad_request(self):
        """Test that DELETE method is not allowed for reactivate."""
        response = self.global_client.delete(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class UserReadViewTestCase(UserViewsBaseTestCase):
    """Test case for the User views with API_ACTION_READ action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.read_url = reverse("users:user:read", args=[self.test_user.pk])
        self.read_other_campaign_url = reverse(
            "users:user:read", args=[self.test_other_campaign_user.pk]
        )

    def test_get_read_global_user(self):
        """User with SCOPE_GLOBAL can read a user."""
        response = self.global_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.test_user.username)
        # Check that form fields are disabled
        self.assertContains(response, "disabled")

    def test_get_read_global_user_other_campaign(self):
        """User with SCOPE_GLOBAL can read a user from another campaign."""
        response = self.global_client.get(self.read_other_campaign_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.test_other_campaign_user.username)
        # Check that form fields are disabled
        self.assertContains(response, "disabled")

    def test_get_read_campaign_user_same_campaign(self):
        """User with SCOPE_CAMPAIGN can read a user from the same campaign."""
        response = self.campaign_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.test_user.username)
        # Check that form fields are disabled
        self.assertContains(response, "disabled")

    def test_get_read_campaign_user_other_campaign_forbidden(self):
        """User with SCOPE_CAMPAIGN cannot read a user from another campaign."""
        response = self.campaign_client.get(self.read_other_campaign_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_read_list_only_user(self):
        """User with only list permission can read a user."""
        response = self.list_only_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.test_user.username)
        # Check that form fields are disabled
        self.assertContains(response, "disabled")

    def test_get_read_none_user_forbidden(self):
        """User with SCOPE_NONE cannot read a user."""
        response = self.none_client.get(self.read_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_read_anonymous_user_redirect(self):
        """Test that an anonymous user cannot read a user."""
        response = self.anonymous_client.get(self.read_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_read_bad_request(self):
        """Test that POST method is not allowed for read."""
        response = self.global_client.post(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_read_bad_request(self):
        """Test that DELETE method is not allowed for read."""
        response = self.global_client.delete(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class UserHistoryViewTestCase(UserViewsBaseTestCase):
    """Test case for the User views with API_ACTION_HISTORY action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.history_url = reverse("users:user:history", args=[self.test_user.pk])
        self.history_other_campaign_url = reverse(
            "users:user:history", args=[self.test_other_campaign_user.pk]
        )

    def test_get_history_global_user(self):
        """User with SCOPE_GLOBAL can view history of a user."""
        response = self.global_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion

    def test_get_history_global_user_other_campaign(self):
        """User with SCOPE_GLOBAL can view history of a user from another campaign."""
        response = self.global_client.get(self.history_other_campaign_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion

    def test_get_history_campaign_user_same_campaign(self):
        """User with SCOPE_CAMPAIGN can view history of a user from the same campaign."""
        response = self.campaign_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion

    def test_get_history_campaign_user_other_campaign_forbidden(self):
        """User with SCOPE_CAMPAIGN cannot view history of a user from another campaign."""
        response = self.campaign_client.get(self.history_other_campaign_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_history_list_only_user(self):
        """User with only list permission can view history of a user."""
        response = self.list_only_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion

    def test_get_history_none_user_forbidden(self):
        """User with SCOPE_NONE cannot view history of a user."""
        response = self.none_client.get(self.history_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_history_anonymous_user_redirect(self):
        """Test that an anonymous user cannot view history of a user."""
        response = self.anonymous_client.get(self.history_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_history_bad_request(self):
        """Test that POST method is not allowed for history."""
        response = self.global_client.post(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_history_bad_request(self):
        """Test that DELETE method is not allowed for history."""
        response = self.global_client.delete(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class UserListViewTestCase(UserViewsBaseTestCase):
    """Test case for the User views with API_ACTION_LIST action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.list_url = reverse("users:user:list")

    def test_get_list_global_user(self):
        """User with SCOPE_GLOBAL can list all users."""
        response = self.global_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the response contains a table
        self.assertContains(response, "<table")
        # Check that users from both campaigns are listed
        self.assertContains(response, self.test_user.username)
        self.assertContains(response, self.test_other_campaign_user.username)

    def test_get_list_campaign_user(self):
        """User with SCOPE_CAMPAIGN can list users from their campaign."""
        response = self.campaign_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the response contains a table
        self.assertContains(response, "<table")
        # Check that only users from the same campaign are listed
        self.assertContains(response, self.test_user.username)
        self.assertNotContains(response, self.test_other_campaign_user.username)

    def test_get_list_list_only_user(self):
        """User with only list permission can list users."""
        response = self.list_only_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the response contains a table
        self.assertContains(response, "<table")
        # Check that users from both campaigns are listed (since list_only_user has SCOPE_GLOBAL)
        self.assertContains(response, self.test_user.username)
        self.assertContains(response, self.test_other_campaign_user.username)

    def test_get_list_none_user_forbidden(self):
        """User with SCOPE_NONE cannot list users."""
        response = self.none_client.get(self.list_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_list_anonymous_user_redirect(self):
        """Test that an anonymous user cannot list users."""
        response = self.anonymous_client.get(self.list_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_list_bad_request(self):
        """Test that POST method is not allowed for list."""
        response = self.global_client.post(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_list_bad_request(self):
        """Test that DELETE method is not allowed for list."""
        response = self.global_client.delete(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class UserResetViewTestCase(UserViewsBaseTestCase):
    """Test case for the User views with API_ACTION_RESET action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.reset_url = reverse("users:user:reset", args=[self.test_user.pk])
        self.reset_other_campaign_url = reverse(
            "users:user:reset", args=[self.test_other_campaign_user.pk]
        )

    def test_get_reset_global_user(self):
        """User with SCOPE_GLOBAL can access the reset form."""
        response = self.global_client.get(self.reset_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, "password")

    def test_get_reset_global_user_other_campaign(self):
        """User with SCOPE_GLOBAL can reset a user from another campaign."""
        response = self.global_client.get(self.reset_other_campaign_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, "password")

    def test_get_reset_campaign_user_same_campaign(self):
        """User with SCOPE_CAMPAIGN can reset a user from the same campaign."""
        response = self.campaign_client.get(self.reset_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, "password")

    def test_get_reset_campaign_user_other_campaign_forbidden(self):
        """User with SCOPE_CAMPAIGN cannot reset a user from another campaign."""
        response = self.campaign_client.get(self.reset_other_campaign_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_reset_none_user_forbidden(self):
        """User with SCOPE_NONE cannot access the reset form."""
        response = self.none_client.get(self.reset_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_reset_list_only_user_forbidden(self):
        """User with only list permission cannot access the reset form."""
        response = self.list_only_client.get(self.reset_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_reset_anonymous_user_redirect(self):
        """Test that an anonymous user cannot access the reset form."""
        response = self.anonymous_client.get(self.reset_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_reset_global_user_success(self):
        """User with SCOPE_GLOBAL can reset a user's password."""
        old_password = self.test_user.password
        data = {"password": "newpassword"}
        response = self.global_client.post(self.reset_url, data)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.test_user.refresh_from_db()
        self.assertNotEqual(self.test_user.password, old_password)

    def test_post_reset_campaign_user_success(self):
        """User with SCOPE_CAMPAIGN can reset a user's password from the same campaign."""
        old_password = self.test_user.password
        data = {"password": "newpassword"}
        response = self.campaign_client.post(self.reset_url, data)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.test_user.refresh_from_db()
        self.assertNotEqual(self.test_user.password, old_password)

    def test_post_reset_campaign_user_other_campaign_forbidden(self):
        """User with SCOPE_CAMPAIGN cannot reset a user's password from another campaign."""
        old_password = self.test_other_campaign_user.password
        data = {"password": "newpassword"}
        response = self.campaign_client.post(self.reset_other_campaign_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.test_other_campaign_user.refresh_from_db()
        self.assertEqual(self.test_other_campaign_user.password, old_password)

    def test_post_reset_global_user_failure(self):
        """User with SCOPE_GLOBAL gets form errors on invalid data."""
        old_password = self.test_user.password
        data = {"password": ""}  # Empty password should fail validation
        response = self.global_client.post(self.reset_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Refresh from database to verify no changes
        self.test_user.refresh_from_db()
        self.assertEqual(self.test_user.password, old_password)

    def test_delete_reset_bad_request(self):
        """Test that DELETE method is not allowed for reset."""
        response = self.global_client.delete(self.reset_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class UserExportViewTestCase(UserViewsBaseTestCase):
    """Test case for the User views with API_ACTION_EXPORT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.export_url = reverse("users:user:export")

    def test_get_export_global_user(self):
        """User with SCOPE_GLOBAL can export users."""
        response = self.global_client.get(self.export_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment; filename=", response.get("Content-Disposition", ""))

        # check for items in excel file
        try:
            file_like_object = io.BytesIO(response.content)
            workbook = openpyxl.load_workbook(file_like_object)
        except Exception as e:
            self.fail(f"Error reading excel file: {e}")
        else:
            sheet = workbook.active
            self.assertGreater(sheet.max_row, 1, "No data in excel file")
            users_id = list()
            for row_index in range(2, sheet.max_row + 1):
                row = sheet[row_index]
                users_id.append(row[0].value)  # row[0] first column => id column
            self.assertTrue(User.objects.filter(pk__in=users_id).exists())

    def test_get_export_campaign_user(self):
        """User with SCOPE_CAMPAIGN can export users from their campaign."""
        response = self.campaign_client.get(self.export_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment; filename=", response.get("Content-Disposition", ""))

        # check for items in excel file
        try:
            file_like_object = io.BytesIO(response.content)
            workbook = openpyxl.load_workbook(file_like_object)
        except Exception as e:
            self.fail(f"Error reading excel file: {e}")
        else:
            sheet = workbook.active
            self.assertGreater(sheet.max_row, 1, "No data in excel file")
            users_id = list()
            for row_index in range(2, sheet.max_row + 1):
                row = sheet[row_index]
                users_id.append(row[0].value)  # row[0] first column => id column
            users = User.objects.filter(pk__in=users_id)
            self.assertTrue(bool(users))
            for user in users:
                self.assertEqual(user.campaign, self.test_campaign)

    def test_get_export_none_user_forbidden(self):
        """User with SCOPE_NONE cannot export users."""
        response = self.none_client.get(self.export_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_list_only_user_forbidden(self):
        """User with only list permission cannot export users."""
        response = self.list_only_client.get(self.export_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_anonymous_user_redirect(self):
        """Test that an anonymous user cannot export users."""
        response = self.anonymous_client.get(self.export_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_export_bad_request(self):
        """Test that POST method is not allowed for export."""
        response = self.global_client.post(self.export_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_export_bad_request(self):
        """Test that DELETE method is not allowed for export."""
        response = self.global_client.delete(self.export_url)
        self.assertEqual(response.status_code, 400)  # Bad request method
