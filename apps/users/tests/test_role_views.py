"""
Tests for the Role views with various API actions.
"""

import io

from django.contrib.auth.hashers import make_password
from django.test import Client, TestCase
from django.urls import reverse

import openpyxl

from apps.common.constants import SCOPE_CAMPAIGN, SCOPE_GLOBAL
from apps.users.models import Campaign, Role, User


class RoleViewsBaseTestCase(TestCase):
    """Base test case for Role views with various API actions."""

    fixtures = ["prod/roles.json", "prod/campaigns.json", "prod/users.json"]

    def setUp(self):
        """Set up test data."""
        # Create a test campaign
        self.test_campaign = Campaign.objects.create(
            name="TEST CAMPAIGN", description="Campaign for testing", is_active=True
        )

        # Create roles with different permissions
        self.admin_role = Role.objects.create(
            name="TEST ADMIN ROLE",
            is_active=True,
            can_add_role=True,
            can_edit_role=True,
            can_delete_role=True,
            can_export_role=True,
            can_history_role=True,
        )

        self.list_only_role = Role.objects.create(
            name="TEST LIST ONLY ROLE",
            is_active=True,
            can_add_role=False,
            can_edit_role=False,
            can_delete_role=False,
            can_export_role=False,
            can_history_role=True,  # This enables can_list_role
        )

        self.no_perm_role = Role.objects.create(
            name="TEST NO PERM ROLE",
            is_active=True,
            can_add_role=False,
            can_edit_role=False,
            can_delete_role=False,
            can_export_role=False,
            can_history_role=False,
        )

        # Create users with different roles
        self.admin_user = User.objects.create(
            username="99999999",
            password=make_password("password"),
            first_name="ADMIN",
            last_name="USER",
            document_type=1,
            document_number="99999999",
            role=self.admin_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.list_only_user = User.objects.create(
            username="88888888",
            password=make_password("password"),
            first_name="LIST ONLY",
            last_name="USER",
            document_type=1,
            document_number="88888888",
            role=self.list_only_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.no_perm_user = User.objects.create(
            username="77777777",
            password=make_password("password"),
            first_name="NO PERM",
            last_name="USER",
            document_type=1,
            document_number="77777777",
            role=self.no_perm_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a test role that will be the target of actions
        self.test_role = Role.objects.create(
            name="TEST ROLE",
            is_active=True,
            can_add_role=False,
            can_edit_role=False,
            can_delete_role=False,
            can_export_role=False,
            can_history_role=False,
        )

        # Create a soft-deleted role for reactivation tests
        self.deleted_role = Role.objects.create(
            name="DELETED ROLE",
            is_active=False,
            can_add_role=False,
            can_edit_role=False,
            can_delete_role=False,
            can_export_role=False,
            can_history_role=False,
        )

        # Create clients for each user
        self.admin_client = Client()
        self.list_only_client = Client()
        self.no_perm_client = Client()
        self.anonymous_client = Client()

        # Log in the clients
        self.admin_client.login(username="99999999", password="password")
        self.list_only_client.login(username="88888888", password="password")
        self.no_perm_client.login(username="77777777", password="password")


class RoleAddViewTestCase(RoleViewsBaseTestCase):
    """Test case for the Role views with API_ACTION_ADD action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.add_url = reverse("users:role:add")

    def test_get_add_admin_user(self):
        """User with can_add_role permission can access the add form."""
        response = self.admin_client.get(self.add_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")

    def test_get_add_list_only_user(self):
        """User with only list permission cannot access the add form."""
        response = self.list_only_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_no_perm_user(self):
        """User with no permissions cannot access the add form."""
        response = self.no_perm_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_anonymous_user(self):
        """Test that an anonymous user cannot access the add form."""
        response = self.anonymous_client.get(self.add_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_add_admin_user_success(self):
        """User with can_add_role permission can access the add form."""
        data = {
            "name": "NEW ROLE",
            "scope_processresult": SCOPE_CAMPAIGN,  # SCOPE_CAMPAIGN
            "scope_process": SCOPE_CAMPAIGN,  # SCOPE_CAMPAIGN
            "scope_typification": SCOPE_CAMPAIGN,  # SCOPE_CAMPAIGN
            "scope_wordlist": SCOPE_CAMPAIGN,  # SCOPE_CAMPAIGN
            "scope_agent": SCOPE_CAMPAIGN,  # SCOPE_CAMPAIGN
            "scope_user": SCOPE_GLOBAL,  # SCOPE_GLOBAL
            "can_add_role": "on",
            "can_edit_role": "on",
            "can_delete_role": "on",
            "can_export_role": "on",
            "can_history_role": "on",
        }
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify user was created
        new_role = Role.objects.filter(name="NEW ROLE").first()
        self.assertIsNotNone(new_role)
        self.assertEqual(new_role.scope_processresult, SCOPE_CAMPAIGN)
        self.assertEqual(new_role.scope_process, SCOPE_CAMPAIGN)
        self.assertEqual(new_role.scope_typification, SCOPE_CAMPAIGN)
        self.assertEqual(new_role.scope_wordlist, SCOPE_CAMPAIGN)
        self.assertEqual(new_role.scope_agent, SCOPE_CAMPAIGN)
        self.assertEqual(new_role.scope_user, SCOPE_GLOBAL)
        self.assertTrue(new_role.can_add_role)
        self.assertTrue(new_role.can_edit_role)
        self.assertTrue(new_role.can_delete_role)
        self.assertTrue(new_role.can_export_role)
        self.assertTrue(new_role.can_history_role)
        self.assertTrue(new_role.is_active)

    def test_post_add_admin_user_failure(self):
        """User with can_add_role permission gets form errors on invalid data."""
        data = {
            "name": "",  # Empty name should fail validation
            "can_add_role": True,
            "can_edit_role": True,
            "can_delete_role": True,
            "can_export_role": True,
            "can_history_role": True,
        }
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify role was not created
        self.assertFalse(Role.objects.filter(name="").exists())

    def test_post_add_list_only_user_forbidden(self):
        """User with only list permission cannot add a new role."""
        data = {
            "name": "SHOULD NOT BE CREATED",
            "can_add_role": True,
            "can_edit_role": True,
            "can_delete_role": True,
            "can_export_role": True,
            "can_history_role": True,
        }
        response = self.list_only_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify role was not created
        self.assertFalse(Role.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_post_add_no_perm_user_forbidden(self):
        """User with no permissions cannot add a new role."""
        data = {
            "name": "SHOULD NOT BE CREATED",
            "can_add_role": True,
            "can_edit_role": True,
            "can_delete_role": True,
            "can_export_role": True,
            "can_history_role": True,
        }
        response = self.no_perm_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify role was not created
        self.assertFalse(Role.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_post_add_anonymous_user_redirect(self):
        """Test that an anonymous user cannot add a new role."""
        data = {
            "name": "SHOULD NOT BE CREATED",
            "can_add_role": True,
            "can_edit_role": True,
            "can_delete_role": True,
            "can_export_role": True,
            "can_history_role": True,
        }
        response = self.anonymous_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify role was not created
        self.assertFalse(Role.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_delete_add_bad_request(self):
        """Test that DELETE method is not allowed for add."""
        response = self.admin_client.delete(self.add_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class RoleHomeViewTestCase(RoleViewsBaseTestCase):
    """Test case for the Role views with API_ACTION_HOME action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.home_url = reverse("users:role:home")

    def test_get_home_admin_user(self):
        """User with can_list_role permission can access the home page."""
        response = self.admin_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        # Validate template used
        self.assertTemplateUsed(response, "users/role/home.html")

    def test_get_home_list_only_user(self):
        """User with only list permission can access the home page."""
        response = self.list_only_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "users/role/home.html")

    def test_get_home_no_perm_user(self):
        """User with no permissions cannot access the home page."""
        response = self.no_perm_client.get(self.home_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_home_anonymous_user(self):
        """Test that an anonymous user cannot access the home page."""
        response = self.anonymous_client.get(self.home_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_home_bad_request(self):
        """Test that POST method is not allowed for home page."""
        response = self.admin_client.post(self.home_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class RoleEditViewTestCase(RoleViewsBaseTestCase):
    """Test case for the Role views with API_ACTION_EDIT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.edit_url = reverse("users:role:edit", args=[self.test_role.pk])

    def test_get_edit_admin_user(self):
        """User with can_edit_role permission can access the edit form."""
        response = self.admin_client.get(self.edit_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, self.test_role.name)

    def test_get_edit_list_only_user(self):
        """User with only list permission cannot access the edit form."""
        response = self.list_only_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_no_perm_user(self):
        """User with no permissions cannot access the edit form."""
        response = self.no_perm_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_anonymous_user(self):
        """Test that an anonymous user cannot access the edit form."""
        response = self.anonymous_client.get(self.edit_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_edit_admin_user_success(self):
        """User with can_edit_role permission can access the edit form."""
        data = {
            "name": "UPDATED ROLE",
            "scope_processresult": SCOPE_CAMPAIGN,  # SCOPE_CAMPAIGN
            "scope_process": SCOPE_CAMPAIGN,  # SCOPE_CAMPAIGN
            "scope_typification": SCOPE_CAMPAIGN,  # SCOPE_CAMPAIGN
            "scope_wordlist": SCOPE_CAMPAIGN,  # SCOPE_CAMPAIGN
            "scope_agent": SCOPE_CAMPAIGN,  # SCOPE_CAMPAIGN
            "scope_user": SCOPE_GLOBAL,  # SCOPE_GLOBAL
            "can_add_role": "on",
            "can_edit_role": "on",
            "can_delete_role": "on",
            "can_export_role": "on",
            "can_history_role": "on",
        }
        original_pk = self.test_role.pk
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Verify user was created
        new_role = Role.objects.filter(name="UPDATED ROLE").first()
        self.assertIsNotNone(new_role)
        self.assertEqual(original_pk, new_role.pk)
        self.assertEqual(new_role.scope_processresult, SCOPE_CAMPAIGN)
        self.assertEqual(new_role.scope_process, SCOPE_CAMPAIGN)
        self.assertEqual(new_role.scope_typification, SCOPE_CAMPAIGN)
        self.assertEqual(new_role.scope_wordlist, SCOPE_CAMPAIGN)
        self.assertEqual(new_role.scope_agent, SCOPE_CAMPAIGN)
        self.assertEqual(new_role.scope_user, SCOPE_GLOBAL)
        self.assertTrue(new_role.can_add_role)
        self.assertTrue(new_role.can_edit_role)
        self.assertTrue(new_role.can_delete_role)
        self.assertTrue(new_role.can_export_role)
        self.assertTrue(new_role.can_history_role)
        self.assertTrue(new_role.is_active)

    def test_post_edit_admin_user_failure(self):
        """User with can_edit_role permission gets form errors on invalid data."""
        data = {
            "can_add_role": True,
            "can_edit_role": True,
            "can_delete_role": True,
            "can_export_role": True,
            "can_history_role": True,
            "scope_wordlist": SCOPE_GLOBAL,  # GLOBAL SCOPE IS NOT ALLOWED FOR WORDLIST
        }
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Refresh from database to verify no changes
        self.test_role.refresh_from_db()
        self.assertNotEqual(self.test_role.name, "")
        self.assertNotEqual(self.test_role.scope_wordlist, SCOPE_GLOBAL)

    def test_post_edit_list_only_user_forbidden(self):
        """User with only list permission cannot edit a role."""
        data = {
            "name": "SHOULD NOT UPDATE",
            "can_add_role": True,
            "can_edit_role": True,
            "can_delete_role": True,
            "can_export_role": True,
            "can_history_role": True,
        }
        response = self.list_only_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.test_role.refresh_from_db()
        self.assertNotEqual(self.test_role.name, "SHOULD NOT UPDATE")

    def test_post_edit_no_perm_user_forbidden(self):
        """User with no permissions cannot edit a role."""
        data = {
            "name": "SHOULD NOT UPDATE",
            "can_add_role": True,
            "can_edit_role": True,
            "can_delete_role": True,
            "can_export_role": True,
            "can_history_role": True,
        }
        response = self.no_perm_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.test_role.refresh_from_db()
        self.assertNotEqual(self.test_role.name, "SHOULD NOT UPDATE")

    def test_post_edit_anonymous_user_redirect(self):
        """Test that an anonymous user cannot edit a role."""
        data = {
            "name": "SHOULD NOT UPDATE",
            "can_add_role": True,
            "can_edit_role": True,
            "can_delete_role": True,
            "can_export_role": True,
            "can_history_role": True,
        }
        response = self.anonymous_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.test_role.refresh_from_db()
        self.assertNotEqual(self.test_role.name, "SHOULD NOT UPDATE")

    def test_delete_edit_bad_request(self):
        """Test that DELETE method is not allowed for edit."""
        response = self.admin_client.delete(self.edit_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class RoleDeleteViewTestCase(RoleViewsBaseTestCase):
    """Test case for the Role views with API_ACTION_DELETE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.delete_url = reverse("users:role:delete", args=[self.test_role.pk])

    def test_delete_admin_user_success(self):
        """User with can_delete_role permission can soft-delete a role."""
        response = self.admin_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.test_role.refresh_from_db()
        self.assertFalse(self.test_role.is_active)

    def test_delete_list_only_user_forbidden(self):
        """User with only list permission cannot soft-delete a role."""
        response = self.list_only_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.test_role.refresh_from_db()
        self.assertTrue(self.test_role.is_active)

    def test_delete_no_perm_user_forbidden(self):
        """User with no permissions cannot soft-delete a role."""
        response = self.no_perm_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.test_role.refresh_from_db()
        self.assertTrue(self.test_role.is_active)

    def test_delete_anonymous_user_redirect(self):
        """Test that an anonymous user cannot soft-delete a role."""
        response = self.anonymous_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.test_role.refresh_from_db()
        self.assertTrue(self.test_role.is_active)

    def test_get_delete_bad_request(self):
        """Test that GET method is not allowed for delete."""
        response = self.admin_client.get(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_post_delete_bad_request(self):
        """Test that POST method is not allowed for delete."""
        response = self.admin_client.post(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class RoleReactivateViewTestCase(RoleViewsBaseTestCase):
    """Test case for the Role views with API_ACTION_REACTIVATE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.reactivate_url = reverse("users:role:reactivate", args=[self.deleted_role.pk])

    def test_post_reactivate_admin_user_success(self):
        """User with can_delete_role permission can reactivate a soft-deleted role."""
        response = self.admin_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.deleted_role.refresh_from_db()
        self.assertTrue(self.deleted_role.is_active)

    def test_post_reactivate_list_only_user_forbidden(self):
        """User with only list permission cannot reactivate a soft-deleted role."""
        response = self.list_only_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.deleted_role.refresh_from_db()
        self.assertFalse(self.deleted_role.is_active)

    def test_post_reactivate_no_perm_user_forbidden(self):
        """User with no permissions cannot reactivate a soft-deleted role."""
        response = self.no_perm_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.deleted_role.refresh_from_db()
        self.assertFalse(self.deleted_role.is_active)

    def test_post_reactivate_anonymous_user_redirect(self):
        """Test that an anonymous user cannot reactivate a soft-deleted role."""
        response = self.anonymous_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.deleted_role.refresh_from_db()
        self.assertFalse(self.deleted_role.is_active)

    def test_get_reactivate_bad_request(self):
        """Test that GET method is not allowed for reactivate."""
        response = self.admin_client.get(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_reactivate_bad_request(self):
        """Test that DELETE method is not allowed for reactivate."""
        response = self.admin_client.delete(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class RoleReadViewTestCase(RoleViewsBaseTestCase):
    """Test case for the Role views with API_ACTION_READ action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.read_url = reverse("users:role:read", args=[self.test_role.pk])

    def test_get_read_admin_user(self):
        """User with can_list_role permission can read a role."""
        response = self.admin_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.test_role.name)
        # Check that form fields are disabled
        self.assertContains(response, "disabled")

    def test_get_read_list_only_user(self):
        """User with only list permission can read a role."""
        response = self.list_only_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.test_role.name)
        # Check that form fields are disabled
        self.assertContains(response, "disabled")

    def test_get_read_no_perm_user_forbidden(self):
        """User with no permissions cannot read a role."""
        response = self.no_perm_client.get(self.read_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_read_anonymous_user_redirect(self):
        """Test that an anonymous user cannot read a role."""
        response = self.anonymous_client.get(self.read_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_read_bad_request(self):
        """Test that POST method is not allowed for read."""
        response = self.admin_client.post(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_read_bad_request(self):
        """Test that DELETE method is not allowed for read."""
        response = self.admin_client.delete(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class RoleHistoryViewTestCase(RoleViewsBaseTestCase):
    """Test case for the Role views with API_ACTION_HISTORY action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.history_url = reverse("users:role:history", args=[self.test_role.pk])

    def test_get_history_admin_user(self):
        """User with can_history_role permission can view history of a role."""
        response = self.admin_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion

    def test_get_history_list_only_user(self):
        """User with only list permission can view history of a role."""
        response = self.list_only_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion

    def test_get_history_no_perm_user_forbidden(self):
        """User with no permissions cannot view history of a role."""
        response = self.no_perm_client.get(self.history_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_history_anonymous_user_redirect(self):
        """Test that an anonymous user cannot view history of a role."""
        response = self.anonymous_client.get(self.history_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_history_bad_request(self):
        """Test that POST method is not allowed for history."""
        response = self.admin_client.post(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_history_bad_request(self):
        """Test that DELETE method is not allowed for history."""
        response = self.admin_client.delete(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class RoleListViewTestCase(RoleViewsBaseTestCase):
    """Test case for the Role views with API_ACTION_LIST action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.list_url = reverse("users:role:list")

    def test_get_list_admin_user(self):
        """User with can_list_role permission can list roles."""
        response = self.admin_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the response contains a table
        self.assertContains(response, "<table")
        # Check that roles are listed
        self.assertContains(response, self.test_role.name)
        self.assertContains(response, self.deleted_role.name)

    def test_get_list_list_only_user(self):
        """User with only list permission can list roles."""
        response = self.list_only_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the response contains a table
        self.assertContains(response, "<table")
        # Check that roles are listed
        self.assertContains(response, self.test_role.name)
        self.assertContains(response, self.deleted_role.name)

    def test_get_list_no_perm_user_forbidden(self):
        """User with no permissions cannot list roles."""
        response = self.no_perm_client.get(self.list_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_list_anonymous_user_redirect(self):
        """Test that an anonymous user cannot list roles."""
        response = self.anonymous_client.get(self.list_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_list_bad_request(self):
        """Test that POST method is not allowed for list."""
        response = self.admin_client.post(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_list_bad_request(self):
        """Test that DELETE method is not allowed for list."""
        response = self.admin_client.delete(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class RoleExportViewTestCase(RoleViewsBaseTestCase):
    """Test case for the Role views with API_ACTION_EXPORT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.export_url = reverse("users:role:export")

    def test_get_export_admin_user(self):
        """User with can_export_role permission can export roles."""
        response = self.admin_client.get(self.export_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment; filename=", response.get("Content-Disposition", ""))

        # check for items in excel file
        try:
            file_like_object = io.BytesIO(response.content)
            workbook = openpyxl.load_workbook(file_like_object)
        except Exception as e:
            self.fail(f"Error reading excel file: {e}")
        else:
            sheet = workbook.active
            self.assertGreater(sheet.max_row, 1, "No data in excel file")
            roles_id = list()
            for row_index in range(2, sheet.max_row + 1):
                row = sheet[row_index]
                roles_id.append(row[0].value)  # row[0] first column => id column
            self.assertTrue(Role.objects.filter(pk__in=roles_id).exists())

    def test_get_export_list_only_user_forbidden(self):
        """User with only list permission cannot export roles."""
        response = self.list_only_client.get(self.export_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_no_perm_user_forbidden(self):
        """User with no permissions cannot export roles."""
        response = self.no_perm_client.get(self.export_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_anonymous_user_redirect(self):
        """Test that an anonymous user cannot export roles."""
        response = self.anonymous_client.get(self.export_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_export_bad_request(self):
        """Test that POST method is not allowed for export."""
        response = self.admin_client.post(self.export_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_export_bad_request(self):
        """Test that DELETE method is not allowed for export."""
        response = self.admin_client.delete(self.export_url)
        self.assertEqual(response.status_code, 400)  # Bad request method
