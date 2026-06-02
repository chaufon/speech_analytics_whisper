"""
Tests for the Campaign views with various API actions.
"""

import io

from django.contrib.auth.hashers import make_password
from django.test import Client, TestCase
from django.urls import reverse

import openpyxl

from apps.users.models import Campaign, Role, User


class CampaignViewsBaseTestCase(TestCase):
    """Base test case for Campaign views with various API actions."""

    fixtures = ["prod/roles.json", "prod/campaigns.json", "prod/users.json"]

    def setUp(self):
        """Set up test data."""
        # Create roles with different permissions
        self.admin_role = Role.objects.create(
            name="TEST ADMIN ROLE",
            is_active=True,
            can_add_campaign=True,
            can_edit_campaign=True,
            can_delete_campaign=True,
            can_export_campaign=True,
            can_history_campaign=True,
        )

        self.list_only_role = Role.objects.create(
            name="TEST LIST ONLY ROLE",
            is_active=True,
            can_add_campaign=False,
            can_edit_campaign=False,
            can_delete_campaign=False,
            can_export_campaign=False,
            can_history_campaign=True,  # This enables can_list_campaign
        )

        self.no_perm_role = Role.objects.create(
            name="TEST NO PERM ROLE",
            is_active=True,
            can_add_campaign=False,
            can_edit_campaign=False,
            can_delete_campaign=False,
            can_export_campaign=False,
            can_history_campaign=False,
        )

        # Create a test campaign
        self.test_campaign = Campaign.objects.create(
            name="TEST CAMPAIGN", description="Campaign for testing", is_active=True
        )

        # Create users with different roles
        self.admin_user = User.objects.create(
            username="99999999",
            password=make_password("password"),
            first_name="ADMIN",
            last_name="USER",
            document_type=1,
            document_number="99999999",
            role=self.admin_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.list_only_user = User.objects.create(
            username="88888888",
            password=make_password("password"),
            first_name="LIST ONLY",
            last_name="USER",
            document_type=1,
            document_number="88888888",
            role=self.list_only_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        self.no_perm_user = User.objects.create(
            username="77777777",
            password=make_password("password"),
            first_name="NO PERM",
            last_name="USER",
            document_type=1,
            document_number="77777777",
            role=self.no_perm_role,
            campaign=self.test_campaign,
            is_active=True,
        )

        # Create a test campaign that will be the target of actions
        self.target_campaign = Campaign.objects.create(
            name="TARGET CAMPAIGN", description="Campaign for testing actions", is_active=True
        )

        # Create a soft-deleted campaign for reactivation tests
        self.deleted_campaign = Campaign.objects.create(
            name="DELETED CAMPAIGN", description="Deleted campaign for testing", is_active=False
        )

        # Create clients for each user
        self.admin_client = Client()
        self.list_only_client = Client()
        self.no_perm_client = Client()
        self.anonymous_client = Client()

        # Log in the clients
        self.admin_client.login(username="99999999", password="password")
        self.list_only_client.login(username="88888888", password="password")
        self.no_perm_client.login(username="77777777", password="password")


class CampaignAddViewTestCase(CampaignViewsBaseTestCase):
    """Test case for the Campaign views with API_ACTION_ADD action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.add_url = reverse("users:campaign:add")

    def test_get_add_admin_user(self):
        """User with can_add_campaign permission can access the add form."""
        response = self.admin_client.get(self.add_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")

    def test_get_add_list_only_user(self):
        """User with only list permission cannot access the add form."""
        response = self.list_only_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_no_perm_user(self):
        """User with no permissions cannot access the add form."""
        response = self.no_perm_client.get(self.add_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_add_anonymous_user(self):
        """Test that an anonymous user cannot access the add form."""
        response = self.anonymous_client.get(self.add_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_add_admin_user_success(self):
        """User with can_add_campaign permission can add a new campaign."""
        data = {"name": "NEW CAMPAIGN", "description": "New campaign for testing"}
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 204)  # Success response

        # Verify campaign was created
        new_campaign = Campaign.objects.filter(name="NEW CAMPAIGN").first()
        self.assertIsNotNone(new_campaign)
        self.assertEqual(new_campaign.description, "New campaign for testing")
        self.assertTrue(new_campaign.is_active)

    def test_post_add_admin_user_failure(self):
        """User with can_add_campaign permission gets form errors on invalid data."""
        data = {"name": "", "description": "Invalid campaign"}  # Empty name should fail validation
        response = self.admin_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Verify campaign was not created
        self.assertFalse(Campaign.objects.filter(description="Invalid campaign").exists())

    def test_post_add_list_only_user_forbidden(self):
        """User with only list permission cannot add a new campaign."""
        data = {"name": "SHOULD NOT BE CREATED", "description": "Should not be created"}
        response = self.list_only_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify campaign was not created
        self.assertFalse(Campaign.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_post_add_no_perm_user_forbidden(self):
        """User with no permissions cannot add a new campaign."""
        data = {"name": "SHOULD NOT BE CREATED", "description": "Should not be created"}
        response = self.no_perm_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Verify campaign was not created
        self.assertFalse(Campaign.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_post_add_anonymous_user_redirect(self):
        """Test that an anonymous user cannot add a new campaign."""
        data = {"name": "SHOULD NOT BE CREATED", "description": "Should not be created"}
        response = self.anonymous_client.post(self.add_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Verify campaign was not created
        self.assertFalse(Campaign.objects.filter(name="SHOULD NOT BE CREATED").exists())

    def test_delete_add_bad_request(self):
        """Test that DELETE method is not allowed for add."""
        response = self.admin_client.delete(self.add_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class CampaignHomeViewTestCase(CampaignViewsBaseTestCase):
    """Test case for the Campaign views with API_ACTION_HOME action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.home_url = reverse("users:campaign:home")

    def test_get_home_admin_user(self):
        """User with can_list_campaign permission can access the home page."""
        response = self.admin_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        # Validate template used
        self.assertTemplateUsed(response, "users/campaign/home.html")

    def test_get_home_list_only_user(self):
        """User with only list permission can access the home page."""
        response = self.list_only_client.get(self.home_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "users/campaign/home.html")

    def test_get_home_no_perm_user(self):
        """User with no permissions cannot access the home page."""
        response = self.no_perm_client.get(self.home_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_home_anonymous_user(self):
        """Test that an anonymous user cannot access the home page."""
        response = self.anonymous_client.get(self.home_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_home_bad_request(self):
        """Test that POST method is not allowed for home page."""
        response = self.admin_client.post(self.home_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class CampaignEditViewTestCase(CampaignViewsBaseTestCase):
    """Test case for the Campaign views with API_ACTION_EDIT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.edit_url = reverse("users:campaign:edit", args=[self.target_campaign.pk])

    def test_get_edit_admin_user(self):
        """User with can_edit_campaign permission can access the edit form."""
        response = self.admin_client.get(self.edit_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "form")
        self.assertContains(response, self.target_campaign.name)

    def test_get_edit_list_only_user(self):
        """User with only list permission cannot access the edit form."""
        response = self.list_only_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_no_perm_user(self):
        """User with no permissions cannot access the edit form."""
        response = self.no_perm_client.get(self.edit_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_edit_anonymous_user(self):
        """Test that an anonymous user cannot access the edit form."""
        response = self.anonymous_client.get(self.edit_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_edit_admin_user_success(self):
        """User with can_edit_campaign permission can edit a campaign."""
        data = {"name": "UPDATED CAMPAIGN", "description": "Updated campaign description"}
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.target_campaign.refresh_from_db()
        self.assertEqual(self.target_campaign.name, "UPDATED CAMPAIGN")
        self.assertEqual(self.target_campaign.description, "Updated campaign description")

    def test_post_edit_admin_user_failure(self):
        """User with can_edit_campaign permission gets form errors on invalid data."""
        data = {"name": "", "description": "Invalid update"}  # Empty name should fail validation
        response = self.admin_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 200)  # Form with errors
        self.assertContains(response, "invalid-feedback")  # Check for validation error class

        # Refresh from database to verify no changes
        self.target_campaign.refresh_from_db()
        self.assertNotEqual(self.target_campaign.name, "")
        self.assertNotEqual(self.target_campaign.description, "Invalid update")

    def test_post_edit_list_only_user_forbidden(self):
        """User with only list permission cannot edit a campaign."""
        data = {"name": "SHOULD NOT UPDATE", "description": "Should not update"}
        response = self.list_only_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.target_campaign.refresh_from_db()
        self.assertNotEqual(self.target_campaign.name, "SHOULD NOT UPDATE")

    def test_post_edit_no_perm_user_forbidden(self):
        """User with no permissions cannot edit a campaign."""
        data = {"name": "SHOULD NOT UPDATE", "description": "Should not update"}
        response = self.no_perm_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.target_campaign.refresh_from_db()
        self.assertNotEqual(self.target_campaign.name, "SHOULD NOT UPDATE")

    def test_post_edit_anonymous_user_redirect(self):
        """Test that an anonymous user cannot edit a campaign."""
        data = {"name": "SHOULD NOT UPDATE", "description": "Should not update"}
        response = self.anonymous_client.post(self.edit_url, data)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.target_campaign.refresh_from_db()
        self.assertNotEqual(self.target_campaign.name, "SHOULD NOT UPDATE")

    def test_delete_edit_bad_request(self):
        """Test that DELETE method is not allowed for edit."""
        response = self.admin_client.delete(self.edit_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class CampaignDeleteViewTestCase(CampaignViewsBaseTestCase):
    """Test case for the Campaign views with API_ACTION_DELETE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.delete_url = reverse("users:campaign:delete", args=[self.target_campaign.pk])

    def test_delete_admin_user_success(self):
        """User with can_delete_campaign permission can soft-delete a campaign."""
        response = self.admin_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.target_campaign.refresh_from_db()
        self.assertFalse(self.target_campaign.is_active)

    def test_delete_list_only_user_forbidden(self):
        """User with only list permission cannot soft-delete a campaign."""
        response = self.list_only_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.target_campaign.refresh_from_db()
        self.assertTrue(self.target_campaign.is_active)

    def test_delete_no_perm_user_forbidden(self):
        """User with no permissions cannot soft-delete a campaign."""
        response = self.no_perm_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.target_campaign.refresh_from_db()
        self.assertTrue(self.target_campaign.is_active)

    def test_delete_anonymous_user_redirect(self):
        """Test that an anonymous user cannot soft-delete a campaign."""
        response = self.anonymous_client.delete(self.delete_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.target_campaign.refresh_from_db()
        self.assertTrue(self.target_campaign.is_active)

    def test_get_delete_bad_request(self):
        """Test that GET method is not allowed for delete."""
        response = self.admin_client.get(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_post_delete_bad_request(self):
        """Test that POST method is not allowed for delete."""
        response = self.admin_client.post(self.delete_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class CampaignReactivateViewTestCase(CampaignViewsBaseTestCase):
    """Test case for the Campaign views with API_ACTION_REACTIVATE action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.reactivate_url = reverse("users:campaign:reactivate", args=[self.deleted_campaign.pk])

    def test_post_reactivate_admin_user_success(self):
        """User with can_delete_campaign permission can reactivate a soft-deleted campaign."""
        response = self.admin_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 204)  # No content with header event

        # Refresh from database
        self.deleted_campaign.refresh_from_db()
        self.assertTrue(self.deleted_campaign.is_active)

    def test_post_reactivate_list_only_user_forbidden(self):
        """User with only list permission cannot reactivate a soft-deleted campaign."""
        response = self.list_only_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.deleted_campaign.refresh_from_db()
        self.assertFalse(self.deleted_campaign.is_active)

    def test_post_reactivate_no_perm_user_forbidden(self):
        """User with no permissions cannot reactivate a soft-deleted campaign."""
        response = self.no_perm_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

        # Refresh from database to verify no changes
        self.deleted_campaign.refresh_from_db()
        self.assertFalse(self.deleted_campaign.is_active)

    def test_post_reactivate_anonymous_user_redirect(self):
        """Test that an anonymous user cannot reactivate a soft-deleted campaign."""
        response = self.anonymous_client.post(self.reactivate_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

        # Refresh from database to verify no changes
        self.deleted_campaign.refresh_from_db()
        self.assertFalse(self.deleted_campaign.is_active)

    def test_get_reactivate_bad_request(self):
        """Test that GET method is not allowed for reactivate."""
        response = self.admin_client.get(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_reactivate_bad_request(self):
        """Test that DELETE method is not allowed for reactivate."""
        response = self.admin_client.delete(self.reactivate_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class CampaignReadViewTestCase(CampaignViewsBaseTestCase):
    """Test case for the Campaign views with API_ACTION_READ action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.read_url = reverse("users:campaign:read", args=[self.target_campaign.pk])

    def test_get_read_admin_user(self):
        """User with can_list_campaign permission can read a campaign."""
        response = self.admin_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_campaign.name)
        # Check that form fields are disabled
        self.assertContains(response, "disabled")

    def test_get_read_list_only_user(self):
        """User with only list permission can read a campaign."""
        response = self.list_only_client.get(self.read_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.target_campaign.name)
        # Check that form fields are disabled
        self.assertContains(response, "disabled")

    def test_get_read_no_perm_user_forbidden(self):
        """User with no permissions cannot read a campaign."""
        response = self.no_perm_client.get(self.read_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_read_anonymous_user_redirect(self):
        """Test that an anonymous user cannot read a campaign."""
        response = self.anonymous_client.get(self.read_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_read_bad_request(self):
        """Test that POST method is not allowed for read."""
        response = self.admin_client.post(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_read_bad_request(self):
        """Test that DELETE method is not allowed for read."""
        response = self.admin_client.delete(self.read_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class CampaignHistoryViewTestCase(CampaignViewsBaseTestCase):
    """Test case for the Campaign views with API_ACTION_HISTORY action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.history_url = reverse("users:campaign:history", args=[self.target_campaign.pk])

    def test_get_history_admin_user(self):
        """User with can_history_campaign permission can view history of a campaign."""
        response = self.admin_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion

    def test_get_history_list_only_user(self):
        """User with only list permission can view history of a campaign."""
        response = self.list_only_client.get(self.history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "accordion")  # Check for accordion

    def test_get_history_no_perm_user_forbidden(self):
        """User with no permissions cannot view history of a campaign."""
        response = self.no_perm_client.get(self.history_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_history_anonymous_user_redirect(self):
        """Test that an anonymous user cannot view history of a campaign."""
        response = self.anonymous_client.get(self.history_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_history_bad_request(self):
        """Test that POST method is not allowed for history."""
        response = self.admin_client.post(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_history_bad_request(self):
        """Test that DELETE method is not allowed for history."""
        response = self.admin_client.delete(self.history_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class CampaignListViewTestCase(CampaignViewsBaseTestCase):
    """Test case for the Campaign views with API_ACTION_LIST action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.list_url = reverse("users:campaign:list")

    def test_get_list_admin_user(self):
        """User with can_list_campaign permission can list campaigns."""
        response = self.admin_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the response contains a table
        self.assertContains(response, "<table")
        # Check that campaigns are listed
        self.assertContains(response, self.target_campaign.name)
        self.assertContains(response, self.deleted_campaign.name)

    def test_get_list_list_only_user(self):
        """User with only list permission can list campaigns."""
        response = self.list_only_client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        # Check that the response contains a table
        self.assertContains(response, "<table")
        # Check that campaigns are listed
        self.assertContains(response, self.target_campaign.name)
        self.assertContains(response, self.deleted_campaign.name)

    def test_get_list_no_perm_user_forbidden(self):
        """User with no permissions cannot list campaigns."""
        response = self.no_perm_client.get(self.list_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_list_anonymous_user_redirect(self):
        """Test that an anonymous user cannot list campaigns."""
        response = self.anonymous_client.get(self.list_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_list_bad_request(self):
        """Test that POST method is not allowed for list."""
        response = self.admin_client.post(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method

    def test_delete_list_bad_request(self):
        """Test that DELETE method is not allowed for list."""
        response = self.admin_client.delete(self.list_url)
        self.assertEqual(response.status_code, 400)  # Bad request method


class CampaignExportViewTestCase(CampaignViewsBaseTestCase):
    """Test case for the Campaign views with API_ACTION_EXPORT action."""

    def setUp(self):
        """Set up test data."""
        super().setUp()
        self.export_url = reverse("users:campaign:export")

    def test_get_export_admin_user(self):
        """User with can_export_campaign permission can export campaigns."""
        response = self.admin_client.get(self.export_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment; filename=", response.get("Content-Disposition", ""))

        # check for items in excel file
        try:
            file_like_object = io.BytesIO(response.content)
            workbook = openpyxl.load_workbook(file_like_object)
        except Exception as e:
            self.fail(f"Error reading excel file: {e}")
        else:
            sheet = workbook.active
            self.assertGreater(sheet.max_row, 1, "No data in excel file")
            campaigns_id = list()
            for row_index in range(2, sheet.max_row + 1):
                row = sheet[row_index]
                campaigns_id.append(row[0].value)  # row[0] first column => id column
            self.assertTrue(Campaign.objects.filter(pk__in=campaigns_id).exists())

    def test_get_export_list_only_user_forbidden(self):
        """User with only list permission cannot export campaigns."""
        response = self.list_only_client.get(self.export_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_no_perm_user_forbidden(self):
        """User with no permissions cannot export campaigns."""
        response = self.no_perm_client.get(self.export_url)
        self.assertEqual(response.status_code, 403)  # Forbidden

    def test_get_export_anonymous_user_redirect(self):
        """Test that an anonymous user cannot export campaigns."""
        response = self.anonymous_client.get(self.export_url)
        self.assertEqual(response.status_code, 302)  # Redirect to login

    def test_post_export_bad_request(self):
        """Test that POST method is not allowed for export."""
        response = self.admin_client.post(self.export_url)
        self.assertEqual(response.status_code, 400)  # Bad request method
